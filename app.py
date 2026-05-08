# app.py
import streamlit as st
import pandas as pd
import os
from datetime import datetime
from io import BytesIO
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from plotly.subplots import make_subplots
import zipfile
import io
import zipfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
# ============================================================================
# CONFIGURAÇÃO INICIAL
# ============================================================================
st.set_page_config(
    page_title="SINISA",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================================
# ESTILOS CSS OTIMIZADOS
# ============================================================================
CSS_STYLES = """
<style>
    :root {
        --primary-color: #1f4788;
        --secondary-color: #2d5aa8;
        --accent-color: #0066cc;
        --light-bg: #f5f7fa;
        --border-color: #d0d8e0;
        --white: #ffffff;
        --dark-text: #333;
        --light-text: #666;
    }
    .appViewContainer, .stAppHeader, .block-container { padding-top: 0 !important; margin-top: 0 !important; }
    header { display: none !important; }
    [data-testid="stTabs"] [role="tablist"] { display: flex; align-items: center; gap: 0 !important; }
    [data-testid="stTabs"] [role="tablist"] button { 
        display: inline-flex; align-items: center; justify-content: center; gap: 0.4rem; 
        padding: 0.6rem 1rem !important; height: 2.5rem; 
    }
    [data-testid="stTabs"] [role="tablist"] button svg { width: 1.2rem; height: 1.2rem; }
    body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: var(--white); margin: 0 !important; padding: 0 !important; }
    .header-container { 
        background: linear-gradient(135deg, #0B3040 0%, #114157 100%); 
        padding: 2rem; border-radius: 8px; margin-bottom: 1rem; 
        color: white; box-shadow: 0 4px 6px rgba(0,0,0,0.1); 
    }
    .header-title { font-size: 2.5rem; font-weight: 700; margin: 0; color: white; }
    .header-subtitle { font-size: 1rem; margin-top: 0.5rem; opacity: 0.95; color: #e8f0ff; }
    .info-card { 
        background: var(--light-bg); border-left: 4px solid var(--accent-color); 
        padding: 1.5rem; border-radius: 6px; margin: 1rem 0; 
        box-shadow: 0 2px 4px rgba(0,0,0,0.05); 
    }
    .info-card-title { font-size: 1.1rem; font-weight: 600; color: var(--primary-color); margin-bottom: 0.5rem; }
    .info-card-content { font-size: 0.95rem; color: var(--dark-text); line-height: 1.6; }
    .detail-section { 
        background: linear-gradient(135deg, #e8eef5 0%, var(--light-bg) 100%) !important; 
        border: 1px solid var(--border-color) !important; border-radius: 8px !important; 
        padding: 1.5rem !important; margin: 1rem 0 !important; 
        box-shadow: 0 2px 8px rgba(31, 71, 136, 0.08) !important; 
    }
    .detail-label { 
        font-weight: 600; color: var(--primary-color); font-size: 0.9rem; 
        text-transform: uppercase; letter-spacing: 0.5px; margin-top: 1rem; margin-bottom: 0.3rem; 
    }
    .detail-value { 
        color: var(--dark-text); font-size: 0.95rem; padding: 0.5rem 0; 
        border-bottom: 1px solid #c0d9f0; 
    }
    .detail-value-multiline { 
        color: var(--dark-text); font-size: 0.95rem; padding: 0.5rem; 
        border-bottom: 1px solid #c0d9f0; white-space: pre-wrap; word-wrap: break-word; 
        font-family: monospace; margin: 0 !important; line-height: 1.4; display: block; 
    }
    .badge { 
        display: inline-block; padding: 0.3rem 0.8rem; border-radius: 20px; 
        font-size: 0.8rem; font-weight: 600; margin-right: 0.5rem; margin-bottom: 0.5rem; 
    }
    .badge-obrigatorio { background-color: #ffe6e6; color: #c41e3a; }
    .badge-opcional { background-color: #e6f2ff; color: var(--accent-color); }
    .badge-modulo { background-color: #f0e6ff; color: #6b21a8; }
    .table-header { 
        background-color: #0B3040; color: white; padding: 0.3rem 0.75rem !important; 
        font-weight: 600; font-size: 1rem; text-align: left; border-radius: 4px; 
        margin: 0 !important; display: flex; align-items: center; height: var(--table-row-height); line-height: 1.1 !important; 
    }
    .table-row { 
        padding: 0.1rem 0.75rem !important; font-size: 1rem; color: var(--dark-text); 
        margin: 0 !important; line-height: 1.3 !important; height: var(--table-row-height); 
        display: flex; align-items: center; word-wrap: break-word; overflow-wrap: break-word; 
    }
    .table-divider { 
        height: 1px; 
        background-color: #e0e0e0; 
        margin: 0 !important; 
        padding: 0 !important;
        display: flex;
        align-items: center;
    }
    .metric-card { 
        background: linear-gradient(135deg, var(--white) 0%, var(--light-bg) 100%); 
        border: 1px solid var(--border-color); border-radius: 8px; padding: 1rem; 
        text-align: center; box-shadow: 0 2px 8px rgba(31, 71, 136, 0.08); 
    }
    .metric-value { font-size: 2rem; font-weight: 700; color: var(--primary-color); margin: 0.3rem 0; }
    .metric-label { 
        font-size: 0.8rem; color: var(--light-text); text-transform: uppercase; 
        letter-spacing: 0.5px; font-weight: 600; 
    }
    .modal-container { 
        background: #d4e4f7 !important; border-radius: 6px !important; padding: 1.5rem !important; 
        margin: 1rem 0 !important; border: 1px solid #a8c5e8 !important; 
    }
    .modal-container > div { background: transparent !important; }
    /* Módulos - Cores */
    .modulo-agua-header { background-color: var(--primary-color); }
    .modulo-agua-subheader { background-color: var(--secondary-color); }
    .modulo-agua-subform { background-color: #d4e4f7; border-left-color: var(--secondary-color); }
    .modulo-agua-subgrupo { background-color: #e8eef5; border-left-color: var(--secondary-color); }
    .modulo-esgoto-header { background-color: #2d7a4a; }
    .modulo-esgoto-subheader { background-color: #3d9a5a; }
    .modulo-esgoto-subform { background-color: #d4e8d9; border-left-color: #3d9a5a; }
    .modulo-esgoto-subgrupo { background-color: #e8f0e8; border-left-color: #3d9a5a; }
    .modulo-info-complementares-header { background-color: #6b5b95; }
    .modulo-info-complementares-subheader { background-color: #8b7bb8; }
    .modulo-info-complementares-subform { background-color: #e8e4f0; border-left-color: #8b7bb8; }
    .modulo-info-complementares-subgrupo { background-color: #f0eef5; border-left-color: #8b7bb8; }
    [data-testid="stVerticalBlock"], [data-testid="stMarkdownContainer"], 
    [data-testid="stColumn"], .stMarkdown, .main, p, h1, h2, h3, h4, h5, h6, 
    [data-testid="stTabs"], [data-testid="stContainer"] { 
        margin: 0 !important; padding: 0 !important; 
    }
</style>
"""
st.markdown(CSS_STYLES, unsafe_allow_html=True)

# ============================================================================
# FUNÇÕES AUXILIARES - UTILITÁRIOS
# ============================================================================
def hex_to_argb(hex_color):
    """Converte cor hex (#RRGGBB) para aRGB (FFRRGGBB)"""
    hex_color = hex_color.lstrip('#')
    return f"FF{hex_color.upper()}"

def get_color_palette():
    """Retorna paleta de cores centralizada para dashboard e Excel"""
    return {
        'água': {
            'header_color': '#1f4788',
            'subheader_color': '#2d5aa8',
            'subform_color': '#d4e4f7',
            'subgrupo_color': '#e8eef5',
            'border_color': '#2d5aa8',
            'excel_header': '#1F4788',
            'excel_subheader': '#2d5aa8'
        },
        'esgoto': {
            'header_color': '#2d7a4a',
            'subheader_color': '#3d9a5a',
            'subform_color': '#d4e8d9',
            'subgrupo_color': '#e8f0e8',
            'border_color': '#3d9a5a',
            'excel_header': '#2d7a4a',
            'excel_subheader': '#3d9a5a'
        },
        'informações complementares': {
            'header_color': '#6b5b95',
            'subheader_color': '#8b7bb8',
            'subform_color': '#e8e4f0',
            'subgrupo_color': '#f0eef5',
            'border_color': '#8b7bb8',
            'excel_header': '#6b5b95',
            'excel_subheader': '#8b7bb8'
        }
    }

def get_modulo_color(modulo):
    """Retorna cores baseadas no módulo (usa paleta centralizada)"""
    palette = get_color_palette()
    modulo_lower = modulo.lower()
    for chave, cores in palette.items():
        if chave in modulo_lower:
            return cores
    # Padrão se não encontrar
    return palette['água']

# Na função sort_modulos() (linha ~300):
def sort_modulos(modulos_list):
    """Ordena módulos: Água, Esgoto, Informações Complementares"""
    modulos_validos = [str(m).strip() for m in modulos_list if pd.notna(m) and str(m).strip() != 'nan']
    agua = [m for m in modulos_validos if 'água' in m.lower()]
    esgoto = [m for m in modulos_validos if 'esgoto' in m.lower()]
    info_complementares = [m for m in modulos_validos if 'água' not in m.lower() and 'esgoto' not in m.lower()]
    return sorted(agua) + sorted(esgoto) + sorted(info_complementares)

def formatar_nome_com_unidade(nome, unidade):
    """Formata nome com unidade"""
    if pd.isna(unidade) or unidade == "" or unidade == "-":
        return nome
    return f"{nome} ({unidade})"

def formatar_brasileiro(valor):
    """Formata número para padrão brasileiro"""
    if pd.isna(valor) or valor == "":
        return ""
    try:
        num = float(valor)
        return f"{num:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except:
        return str(valor)

def get_col_letter(col_num):
    """Converte número de coluna para letra (1->A, 27->AA)"""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(65 + col_num % 26) + result
        col_num //= 26
    return result

@st.cache_data
def obter_anos_disponiveis():
    """Obtém dinamicamente os anos disponíveis (abas) no arquivo Dados_SINISA.xlsx"""
    file_path = os.path.join(DATA_DIR, "Dados_SINISA.xlsx")
    try:
        # Ler nomes de todas as abas do arquivo
        xls = pd.ExcelFile(file_path)
        sheet_names = xls.sheet_names
        
        # Filtrar apenas abas que são números (anos)
        anos = []
        for sheet in sheet_names:
            try:
                ano = int(sheet)
                anos.append(ano)
            except ValueError:
                # Ignorar abas que não são números
                pass
        
        # Ordenar em ordem decrescente (mais recente primeiro)
        anos_ordenados = sorted(anos, reverse=True)
        
        return anos_ordenados if anos_ordenados else [2026]
    except Exception as e:
        st.error(f"❌ Erro ao obter anos disponíveis: {str(e)}")
        return [2026]
# ============================================================================
# CARREGAMENTO E CACHE DE DADOS
# ============================================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")

@st.cache_data
def load_data():
    """Carrega dados do glossário"""
    file_path = os.path.join(DATA_DIR, "Glossario_SINISA.xlsx")
    df = pd.read_excel(file_path, sheet_name=0)
    colunas_remover = ['Coleta', 'Obrigatória?', 'Status', 'Relatório Gerencial - AGEMS']
    df = df.drop(columns=[col for col in colunas_remover if col in df.columns], errors='ignore')
    df.columns = df.columns.str.strip()
    return df

@st.cache_data
def load_dados_agems(ano):
    """Carrega dados AGEMS da aba específica do ano selecionado"""
    file_path = os.path.join(DATA_DIR, "Dados_SINISA.xlsx")
    try:
        # Lê a aba correspondente ao ano
        df_agems = pd.read_excel(file_path, sheet_name=str(ano))
        df_agems.columns = df_agems.columns.str.strip()
        
        # Filtra apenas as linhas que têm "SIM" na coluna "Relatório Gerencial - AGEMS"
        if 'Relatório Gerencial - AGEMS' in df_agems.columns:
            df_agems = df_agems[
                df_agems['Relatório Gerencial - AGEMS'].astype(str).str.strip().str.upper() == 'SIM'
            ].copy()
        
        # Mapeamento de colunas
        mapa_colunas = {
            'Código IBGE - Município': 'codigo_ibge_municipio',
            'Município': 'municipio',
            'Módulo': 'modulo',
            'Formulário': 'formulario',
            'Subformulário/Grupo': 'subformulario_grupo', 
            'Subgrupo/Palavra-chave': 'subgrupo_palavra_chave',
            'Área Responsável': 'area_responsavel',
            'Código da informação/indicador': 'codigo_da_informacao_indicador',
            'Nome da informação/indicador': 'nome_da_informacao_indicador',
            'Informação/Indicador': 'informacao_indicador',
            'Unidade': 'unidade',
            'Informação Completa': 'informacao_completa'
        }
        
        # Renomear colunas que existem
        for col_original, col_novo in mapa_colunas.items():
            if col_original in df_agems.columns:
                df_agems = df_agems.rename(columns={col_original: col_novo})
        
        return df_agems
    except Exception as e:
        st.error(f"❌ Erro ao carregar dados AGEMS do ano {ano}: {str(e)}")
        return None

def load_dados_agems_por_ano(ano):
    """Carrega dados AGEMS da aba específica do ano"""
    file_path = os.path.join(DATA_DIR, "Dados_SINISA.xlsx")
    try:
        # Lê a aba do ano selecionado
        df_agems = pd.read_excel(file_path, sheet_name=str(ano))
        df_agems.columns = df_agems.columns.str.strip()
        
        # Filtra apenas as linhas que têm "SIM" na coluna "Relatório Gerencial - AGEMS"
        if 'Relatório Gerencial - AGEMS' in df_agems.columns:
            df_agems = df_agems[
                df_agems['Relatório Gerencial - AGEMS'].astype(str).str.strip().str.upper() == 'SIM'
            ].copy()
        
        # Mapeamento de colunas
        mapa_colunas = {
            'Código IBGE - Município': 'codigo_ibge_municipio',
            'Município': 'municipio',
            'Módulo': 'modulo',
            'Formulário': 'formulario',
            'Subformulário/Grupo': 'subformulario_grupo', 
            'Subgrupo/Palavra-chave': 'subgrupo_palavra_chave',
            'Área Responsável': 'area_responsavel',
            'Código da informação/indicador': 'codigo_da_informacao_indicador',
            'Nome da informação/indicador': 'nome_da_informacao_indicador',
            'Informação/Indicador': 'informacao_indicador',
            'Unidade': 'unidade',
            'Informação Completa': 'informacao_completa'
        }
        
        # Renomear colunas que existem
        for col_original, col_novo in mapa_colunas.items():
            if col_original in df_agems.columns:
                df_agems = df_agems.rename(columns={col_original: col_novo})
        
        return df_agems
    except Exception as e:
        st.error(f"❌ Erro ao carregar dados do ano {ano}: {str(e)}")
        return None

@st.cache_data
def load_dados_sinisa(ano=2026):
    """Carrega dados SINISA de uma aba específica (por ano)"""
    file_path = os.path.join(DATA_DIR, "Dados_SINISA.xlsx")
    try:
        # Converter ano para string para usar como nome da aba
        sheet_name = str(ano)
        
        df_sinisa = pd.read_excel(file_path, sheet_name=sheet_name)
        df_sinisa.columns = df_sinisa.columns.str.strip()
        
        # Adicionar coluna de ano para referência
        df_sinisa['ano'] = ano

        mapa_colunas = {
            'Código IBGE - Município': 'codigo_ibge_municipio',
            'Município': 'municipio',
            'Módulo': 'modulo',
            'Formulário': 'formulario',
            'Subformulário/Grupo': 'subformulario_grupo',
            'Subgrupo/Palavra-chave': 'subgrupo_palavra_chave',
            'Área Responsável': 'area_responsavel',
            'Código da informação/indicador': 'codigo_da_informacao_indicador',
            'Nome da informação/indicador': 'nome_da_informacao_indicador',
            'Informação/Indicador': 'informacao_indicador',
            'Unidade': 'unidade',
            'Informação Completa': 'informacao_completa',
            'Anual': 'anual',           
            'Acumulado em 12 meses': 'acumulado_12_meses',
            'Média em 12 meses': 'media_12_meses'        
        }
        for col_original, col_novo in mapa_colunas.items():
            if col_original in df_sinisa.columns:
                df_sinisa = df_sinisa.rename(columns={col_original: col_novo})
        return df_sinisa
    except Exception as e:
        st.error(f"❌ Erro ao carregar dados SINISA: {str(e)}")
        return None

try:
    df = load_data()
except FileNotFoundError:
    st.error(f"❌ Arquivo não encontrado em: {os.path.join(DATA_DIR, 'Glossario_SINISA.xlsx')}")
    st.stop()
except Exception as e:
    st.error(f"❌ Erro ao carregar o arquivo: {str(e)}")
    st.stop()

@st.cache_data
def load_correspondencia_sigis_sinisa():
    """Carrega tabela de correspondência SIGIS-SINISA"""
    file_path = os.path.join(DATA_DIR, "Correspondencia_SIGIS_SINISA.xlsx")
    try:
        df = pd.read_excel(file_path, sheet_name=0)
        df.columns = df.columns.str.strip()
        return df
    except Exception as e:
        st.error(f"Erro ao carregar correspondência: {str(e)}")
        return None

@st.cache_data
def load_dados_sigis():
    """Carrega dados SIGIS"""
    file_path = os.path.join(DATA_DIR, "Dados_SIGIS.xlsx")
    try:
        df = pd.read_excel(file_path, sheet_name=0)
        df.columns = df.columns.str.strip()
        return df
    except Exception as e:
        st.error(f"Erro ao carregar dados SIGIS: {str(e)}")
        return None
# ============================================================================
# FUNÇÕES GENÉRICAS DE FILTROS
# ============================================================================

def aplicar_filtros(df, filtros_dict):
    """Aplica múltiplos filtros a um dataframe"""
    df_filtrado = df.copy()
    
    for coluna, valor in filtros_dict.items():
        if valor and valor != "Todos" and coluna in df_filtrado.columns:
            df_filtrado = df_filtrado[df_filtrado[coluna] == valor]
    
    return df_filtrado
# ============================================================================
# FUNÇÕES DE GERAÇÃO DE EXCEL - OTIMIZADAS
# ============================================================================

def criar_tabela_html(df_subgrupo, colunas_meses, largura_cod_desc=350, largura_un=60, largura_mes=60):
    """Cria tabela HTML otimizada para relatórios"""
    html = '<div style="overflow-x: auto; margin: 0 0 0.2rem 0;">'
    html += '<table style="width: 100%; border-collapse: collapse; font-size: 0.85rem; border: none;">'
    html += '<thead><tr style="background-color: #F2F2F2; color: black; position: sticky; top: 0;">'
    html += f'<th style="padding: 0.6rem; text-align: left; border: 1px solid #d0d8e0; width: {largura_cod_desc}px; min-width: {largura_cod_desc}px;">Código - Descrição</th>'
    html += f'<th style="padding: 0.6rem; text-align: center; border: 1px solid #d0d8e0; width: {largura_un}px; min-width: {largura_un}px;">Un.</th>'
    for mes in colunas_meses:
        html += f'<th style="padding: 0.6rem; text-align: center; border: 1px solid #d0d8e0; width: {largura_mes}px; min-width: {largura_mes}px; white-space: nowrap;">{mes}</th>'
    html += '</tr></thead><tbody>'
    for idx, row in df_subgrupo.iterrows():
        html += '<tr style="background-color: #ffffff; border-bottom: 1px solid #d0d8e0;">'
        codigo_descricao = f"{row['codigo_da_informacao_indicador']} - {row['nome_da_informacao_indicador']}"
        html += f'<td style="padding: 0.5rem; text-align: left; border: 1px solid #d0d8e0; width: {largura_cod_desc}px; min-width: {largura_cod_desc}px; font-weight: 500;">{codigo_descricao}</td>'
        html += f'<td style="padding: 0.5rem; text-align: center; border: 1px solid #d0d8e0; width: {largura_un}px; min-width: {largura_un}px;">{row["unidade"]}</td>'
        for mes in colunas_meses:
            valor_formatado = formatar_brasileiro(row[mes])
            html += f'<td style="padding: 0.5rem; text-align: right; border: 1px solid #d0d8e0; width: {largura_mes}px; min-width: {largura_mes}px; font-family: \'Calibri\', monospace; font-weight: 500;">{valor_formatado}</td>'
        html += '</tr>'
    html += '</tbody></table></div>'
    return html

def criar_relatorio_excel_generico(df_para_exportar, municipio_selecionado, colunas_meses):
    """
    Cria relatório Excel genérico para município com estética alinhada ao dashboard
    - Cabeçalho nativo do Excel (aparece apenas na impressão)
    - Tratamento específico para Informações Complementares
    - Paleta de cores consistente
    - Tipografia profissional
    - Bordas e espaçamento otimizados
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    import pandas as pd

    # ====== ESTILOS INLINE ======
    estilos_header_principal = {
        'bg_color': '0B3040',
        'font_color': 'FFFFFF',
        'font_size': 12,
        'bold': True
    }
    estilos_header_secundario = {
        'bg_color': '1F4788',
        'font_color': 'FFFFFF',
        'font_size': 11,
        'bold': True
    }
    estilos_row_alternado = 'F5F7FA'
    estilos_row_normal = 'FFFFFF'
    border_color = 'D0D8E0'
    font_name = 'Calibri'
    font_size_data = 11

    wb = Workbook()
    ws = wb.active
    ws.title = str(municipio_selecionado)[:31]
    ws.sheet_view.zoomScale = 90

    # ====== CABEÇALHO NATIVO DO EXCEL (SINTAXE CORRETA) ======
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'landscape'
    ws.page_margins.left = 1
    ws.page_margins.right = 1
    ws.page_margins.top = 1
    ws.page_margins.bottom = 1
    
    # Definir cabeçalho e rodapé
    ws.oddHeader.left.text = f"Empresa de Saneamento de Mato Grosso do Sul - SANESUL\nRelatório Gerencial - {municipio_selecionado} | {datetime.now().strftime('%m/%Y')} "
    
    ws.oddFooter.right.text = "Página &P de &N"

    # Definir área de impressão até coluna N com 1 página de largura
    ws.print_options.horizontalCentered = False
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToWidth = 1
    ws.print_area = f'A1:N1048576'

    # Definir largura das colunas
    ws.column_dimensions['A'].width = 132
    ws.column_dimensions['B'].width = 17
    for col_num in range(3, 3 + len(colunas_meses)):
        ws.column_dimensions[get_column_letter(col_num)].width = 15

    # ====== COMEÇAR DADOS DIRETO NA LINHA 1 ======
    row_atual = 1

    thin_border = Border(
        left=Side(style='thin', color=border_color),
        right=Side(style='thin', color=border_color),
        top=Side(style='thin', color=border_color),
        bottom=Side(style='thin', color=border_color)
    )

    # ====== CABEÇALHO FIXO DA TABELA (LINHA 1) ======
    headers = ['Código - Descrição', 'Un.'] + colunas_meses
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_atual, column=col_num)
        cell.value = header
        cell.font = Font(name=font_name, size=font_size_data, bold=True, color=estilos_header_principal['font_color'])
        cell.fill = PatternFill(start_color=estilos_header_principal['bg_color'], end_color=estilos_header_principal['bg_color'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[row_atual].height = 20
    
    # Congelar a linha de cabeçalho
    ws.freeze_panes = 'b2'
    row_atual += 1

    # Função auxiliar para imprimir tabela de dados (evita repetição de código)
    def imprimir_tabela_dados(df_dados, row_start, num_cols):
        current_row = row_start
        
        # Dados
        alternado = False
        for idx, row in df_dados.iterrows():
            bg_color = estilos_row_alternado if alternado else estilos_row_normal
            alternado = not alternado
            
            # Coluna 1: Código - Descrição
            cell = ws.cell(row=current_row, column=1)
            cell.value = f"{row['codigo_da_informacao_indicador']} - {row['nome_da_informacao_indicador']}"
            cell.font = Font(name=font_name, size=font_size_data)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
            
            # Coluna 2: Unidade
            cell = ws.cell(row=current_row, column=2)
            cell.value = row['unidade'] if pd.notna(row['unidade']) else '-'
            cell.font = Font(name=font_name, size=font_size_data)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
            
            # Colunas de meses
            for col_num, mes in enumerate(colunas_meses, 3):
                cell = ws.cell(row=current_row, column=col_num)
                valor = row[mes]
                if pd.notna(valor) and str(valor).strip() != "":
                    try:
                        cell.value = float(valor)
                        cell.number_format = '#,##0.00'
                    except:
                        cell.value = str(valor)
                else:
                    cell.value = "-"
                cell.font = Font(name=font_name, size=font_size_data)
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = thin_border
            
            ws.row_dimensions[current_row].height = 18
            current_row += 1
            
        return current_row

    modulos_unicos = sort_modulos(df_para_exportar['modulo'].unique().tolist())
    
    for modulo in modulos_unicos:
        df_modulo = df_para_exportar[df_para_exportar['modulo'] == modulo]
        cores = get_modulo_color(modulo)
        num_cols = 2 + len(colunas_meses)
        
        # ====== CABEÇALHO MÓDULO ======
        ws.merge_cells(f'A{row_atual}:{get_column_letter(num_cols)}{row_atual}')
        cell = ws[f'A{row_atual}']
        cell.value = f" ► {str(modulo).upper()}"
        cell.font = Font(name=font_name, size=estilos_header_principal['font_size'], bold=True, color=estilos_header_principal['font_color'])
        cell.fill = PatternFill(start_color=cores['header_color'].lstrip('#'), end_color=cores['header_color'].lstrip('#'), fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[row_atual].height = 22
        
        for col in range(1, num_cols + 1):
            ws.cell(row=row_atual, column=col).border = thin_border
        row_atual += 1
        
        # ====== VERIFICAÇÃO: INFORMAÇÕES COMPLEMENTARES ======
        if 'informações complementares' in str(modulo).lower():
            # Imprime a tabela direto, sem agrupar por formulário/subformulário
            row_atual = imprimir_tabela_dados(df_modulo, row_atual, num_cols)
            row_atual += 1 # Espaço extra
        else:
            # ====== PROCESSAR FORMULÁRIOS (Água e Esgoto) ======
            for formulario in df_modulo['formulario'].unique():
                if pd.isna(formulario): continue
                df_formulario = df_modulo[df_modulo['formulario'] == formulario]
                
                # Cabeçalho Formulário
                ws.merge_cells(f'A{row_atual}:{get_column_letter(num_cols)}{row_atual}')
                cell = ws[f'A{row_atual}']
                cell.value = f"  {formulario}"
                cell.font = Font(name=font_name, size=estilos_header_secundario['font_size'], bold=True, color=estilos_header_secundario['font_color'])
                cell.fill = PatternFill(start_color=cores['subheader_color'].lstrip('#'), end_color=cores['subheader_color'].lstrip('#'), fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center')
                ws.row_dimensions[row_atual].height = 18
                for col in range(1, num_cols + 1):
                    ws.cell(row=row_atual, column=col).border = thin_border
                row_atual += 1
                
                # ====== PROCESSAR SUBFORMULÁRIOS ======
                for subformulario in df_formulario['subformulario_grupo'].unique():
                    if pd.isna(subformulario): continue
                    df_subformulario = df_formulario[df_formulario['subformulario_grupo'] == subformulario]
                    
                    # Cabeçalho Subformulário
                    ws.merge_cells(f'A{row_atual}:{get_column_letter(num_cols)}{row_atual}')
                    cell = ws[f'A{row_atual}']
                    cell.value = f"    {subformulario}"
                    cell.font = Font(name=font_name, size=10, bold=True, color='1F4788')
                    cell.fill = PatternFill(start_color=cores['subform_color'].lstrip('#'), end_color=cores['subform_color'].lstrip('#'), fill_type='solid')
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    ws.row_dimensions[row_atual].height = 16
                    for col in range(1, num_cols + 1):
                        ws.cell(row=row_atual, column=col).border = thin_border
                    row_atual += 1
                    
                    # ====== PROCESSAR SUBGRUPOS ======
                    for subgrupo in df_subformulario['subgrupo_palavra_chave'].unique():
                        if pd.isna(subgrupo): continue
                        df_subgrupo = df_subformulario[df_subformulario['subgrupo_palavra_chave'] == subgrupo]
                        
                        # Cabeçalho Subgrupo
                        ws.merge_cells(f'A{row_atual}:{get_column_letter(num_cols)}{row_atual}')
                        cell = ws[f'A{row_atual}']
                        cell.value = f"      {subgrupo}"
                        cell.font = Font(name=font_name, size=10, bold=True, color='1F4788')
                        cell.fill = PatternFill(start_color=cores['subgrupo_color'].lstrip('#'), end_color=cores['subgrupo_color'].lstrip('#'), fill_type='solid')
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        ws.row_dimensions[row_atual].height = 15
                        for col in range(1, num_cols + 1):
                            ws.cell(row=row_atual, column=col).border = thin_border
                        row_atual += 1
                        
                        # Imprimir Tabela de Dados
                        row_atual = imprimir_tabela_dados(df_subgrupo, row_atual, num_cols)
                        
                        # Espaço entre subgrupos
                        row_atual += 1
    
    return wb

@st.cache_data
def criar_relatorio_consolidado_excel(df_para_exportar, indicadores_filtrados_tuple, coluna_valor='anual', incluir_colunas_vazias=False, largura_anotacao_min=500):
    """Cria relatório consolidado em Excel com anotações (comentários)
    
    Args:
        df_para_exportar: DataFrame com os dados
        indicadores_filtrados_tuple: Tuple com indicadores filtrados
        coluna_valor: Coluna de valor a usar (padrão: 'anual')
        incluir_colunas_vazias: Se True, inclui colunas vazias
        largura_anotacao_min: Largura mínima da anotação
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
    
    # Carregar glossário para anotações
    df_glossario = load_data()
    
    df_indicadores = pd.DataFrame(indicadores_filtrados_tuple)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"

    # Estilos
    header_fill = PatternFill(start_color='0B3040', end_color='0B3040', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=11)
    thin_border = Border(
        left=Side(style='thin', color='D0D8E0'),
        right=Side(style='thin', color='D0D8E0'),
        top=Side(style='thin', color='D0D8E0'),
        bottom=Side(style='thin', color='D0D8E0')
    )
    
    # Preparar dados
    df_pivot = df_para_exportar.pivot_table(
        index=['codigo_ibge_municipio', 'municipio'],
        columns='codigo_da_informacao_indicador',
        values=coluna_valor,
        aggfunc='first',
        fill_value=None
    ).reset_index()
    
    # Filtrar colunas vazias se necessário
    if not incluir_colunas_vazias:
        colunas_com_dados = ['codigo_ibge_municipio', 'municipio']
        for col in df_pivot.columns:
            if col not in ['codigo_ibge_municipio', 'municipio']:
                tem_dado = False
                for val in df_pivot[col]:
                    if pd.notna(val) and val is not None and str(val).strip() != "" and str(val).strip() != "nan" and str(val).strip() != "None":
                        tem_dado = True
                        break
                if tem_dado:
                    colunas_com_dados.append(col)
        df_pivot = df_pivot[colunas_com_dados]
    
    headers = list(df_pivot.columns)
    
    # Cabeçalhos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        
        # ADICIONAR ANOTAÇÃO NO CABEÇALHO (para colunas de indicadores)
        if col_num > 2:  # Pula código_ibge_municipio e municipio
            codigo_info = header
            row_glossario = df_glossario[df_glossario['Código da informação'] == codigo_info]
            
            if not row_glossario.empty:
                descricao = row_glossario.iloc[0].get('Descrição SINISA', '')
                formula = row_glossario.iloc[0].get('Fórmula', '')
                referencia = row_glossario.iloc[0].get('Referência', '')
                unidade = row_glossario.iloc[0].get('Unidade', '')
                
                # Montar texto da anotação
                texto_anotacao = ""
                
                if pd.notna(descricao) and str(descricao).strip() != "":
                    unidade_texto = f" ({str(unidade).strip()})" if pd.notna(unidade) and str(unidade).strip() != "" else ""
                    texto_anotacao += f"\n\nDESCRIÇÃO COMPLETA: \n{str(descricao).strip()} {unidade_texto}\n\n"
                
                if pd.notna(formula) and str(formula).strip() != "":
                    texto_anotacao += f"FÓRMULA: {str(formula).strip()}\n\n"
                
                if pd.notna(referencia) and str(referencia).strip() != "":
                    texto_anotacao += f"REFERÊNCIA: \n{str(referencia).strip()}"
                
                # Adicionar anotação se houver conteúdo
                if texto_anotacao.strip():
                    comment = Comment(texto_anotacao, "Sistema")
                    
                    # Calcular dimensões
                    num_linhas = texto_anotacao.count('\n') + 1
                    altura_dinamica = max(750, num_linhas * 50)
                    
                    linhas = texto_anotacao.split('\n')
                    linha_mais_longa = max(len(linha) for linha in linhas) if linhas else 0
                    largura_dinamica = max(largura_anotacao_min, linha_mais_longa)
                    
                    comment.width = largura_dinamica
                    comment.height = altura_dinamica
                    comment.moveWith = True
                    
                    cell.comment = comment
    
    # Dados
    for row_num, (idx, row) in enumerate(df_pivot.iterrows(), 2):
        for col_num, col_name in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            value = row[col_name]
            
            if col_num == 1:
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_num == 2:
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                if pd.notna(value) and str(value).strip() != "":
                    try:
                        cell.value = float(value)
                        cell.number_format = '#,##0.00'
                    except:
                        cell.value = value
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
            cell.font = data_font
            cell.border = thin_border
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    for col_num in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 18
    
    # Congelar linhas
    ws.freeze_panes = 'C2'
    
    return wb

@st.cache_data
def criar_exportacao_zip_por_formulario(df_sinisa_tuple, coluna_valor='anual', incluir_colunas_vazias=False, posicao_anotacao_linha=3, largura_anotacao_min=500):
    """Cria ZIP com dados por módulo - Estilo padronizado com o consolidado
    
    Args:
        df_sinisa_tuple: Dados SINISA
        coluna_valor: Coluna de valor a usar (padrão: 'anual')
        incluir_colunas_vazias: Se True, inclui colunas vazias (padrão: False)
        posicao_anotacao_linha: Linha a partir da qual a anotação começa (padrão: 3)
        largura_anotacao_min: Largura mínima da anotação em caracteres (padrão: 50)
    """
    df_sinisa = pd.DataFrame(df_sinisa_tuple)
    zip_buffer = BytesIO()
    
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
    import zipfile
    
    # Carregar glossário para obter Fórmula e Referência
    df_glossario = load_data()
    
    # Estilos baseados no relatório consolidado
    header_fill = PatternFill(start_color='0B3040', end_color='0B3040', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=11)
    thin_border = Border(
        left=Side(style='thin', color='D0D8E0'),
        right=Side(style='thin', color='D0D8E0'),
        top=Side(style='thin', color='D0D8E0'),
        bottom=Side(style='thin', color='D0D8E0')
    )
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for modulo in ['Água', 'Esgoto']:
            df_modulo = df_sinisa[df_sinisa['modulo'] == modulo].copy()
            if len(df_modulo) == 0:
                continue
                
            for info_tipo in ['Informação', 'Indicador']:
                df_tipo = df_modulo[df_modulo['informacao_indicador'] == info_tipo].copy()
                if len(df_tipo) == 0:
                    continue
                    
                wb = Workbook()
                ws_primeiro = wb.active
                ws_primeiro.title = "Índice"
                
                for grupo in sorted(df_tipo['subformulario_grupo'].dropna().unique().tolist()):
                    df_grupo = df_tipo[df_tipo['subformulario_grupo'] == grupo]
                    codigo_nome_unidade = {}
                    
                    for _, row in df_grupo.iterrows():
                        codigo = row['codigo_da_informacao_indicador']
                        codigo_nome_unidade[codigo] = formatar_nome_com_unidade(row['nome_da_informacao_indicador'], row['unidade'])
                    
                    # OBTER LISTA DE TODOS OS INDICADORES DESTE GRUPO (ANTES DO PIVOT)
                    indicadores_do_grupo = sorted(df_grupo['codigo_da_informacao_indicador'].unique().tolist())
                        
                    df_pivot = df_grupo[['codigo_ibge_municipio', 'municipio', 'codigo_da_informacao_indicador', coluna_valor]].copy()
                    df_pivot = df_pivot.drop_duplicates(subset=['codigo_ibge_municipio', 'municipio', 'codigo_da_informacao_indicador'])
                    df_pivot = df_pivot.pivot_table(
                        index=['codigo_ibge_municipio', 'municipio'],
                        columns='codigo_da_informacao_indicador',
                        values=coluna_valor,
                        aggfunc='first',
                        fill_value=None
                    ).reset_index()
                    
                    # REINSERIR COLUNAS QUE O PIVOT TABLE APAGOU POR ESTAREM VAZIAS
                    for indicador in indicadores_do_grupo:
                        if indicador not in df_pivot.columns:
                            df_pivot[indicador] = None
                    
                    # ===== FILTRAR COLUNAS VAZIAS COM VERIFICAÇÃO ROBUSTA =====
                    if not incluir_colunas_vazias:
                        colunas_com_dados = ['codigo_ibge_municipio', 'municipio']
                        for col in df_pivot.columns:
                            if col not in ['codigo_ibge_municipio', 'municipio']:
                                tem_dado = False
                                for val in df_pivot[col]:
                                    # Verificação robusta contra NaN, None e strings "nan"
                                    if pd.notna(val) and val is not None and str(val).strip() != "" and str(val).strip() != "nan" and str(val).strip() != "None":
                                        tem_dado = True
                                        break
                                if tem_dado:
                                    colunas_com_dados.append(col)
                        df_pivot = df_pivot[colunas_com_dados]
                    else:
                        # Se incluir colunas vazias, ordena para ficar na ordem correta
                        colunas_ordenadas = ['codigo_ibge_municipio', 'municipio'] + [col for col in indicadores_do_grupo if col in df_pivot.columns]
                        df_pivot = df_pivot[colunas_ordenadas]
                        
                    ws = wb.create_sheet(title=str(grupo)[:31])
                    headers = list(df_pivot.columns)
                    
                    # Cabeçalhos
                    for col_num, header in enumerate(headers, 1):
                        # Linha 1
                        cell1 = ws.cell(row=1, column=col_num)
                        cell1.value = "" if col_num <= 2 else codigo_nome_unidade.get(header, "")
                        cell1.font = header_font
                        cell1.fill = header_fill
                        cell1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell1.border = thin_border
                        
                        # Linha 2
                        cell2 = ws.cell(row=2, column=col_num)
                        cell2.value = header
                        cell2.font = header_font
                        cell2.fill = header_fill
                        cell2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell2.border = thin_border
                        
                        # ADICIONAR ANOTAÇÃO NA SEGUNDA LINHA
                        if col_num == 1:
                            # Coluna do código - anotação "X"
                            cell2.comment = Comment("X", "Sistema")
                        elif col_num > 2:
                            # Colunas de dados (a partir da 3ª) - buscar Descrição, Fórmula e Referência do glossário
                            codigo_info = header
                            
                            # Buscar no glossário
                            row_glossario = df_glossario[df_glossario['Código da informação'] == codigo_info]
                            
                            if not row_glossario.empty:
                                descricao = row_glossario.iloc[0].get('Descrição SINISA', '')
                                formula = row_glossario.iloc[0].get('Fórmula', '')
                                referencia = row_glossario.iloc[0].get('Referência', '')
                                unidade = row_glossario.iloc[0].get('Unidade', '')
                                area_responsavel = row_glossario.iloc[0].get('Área Responsável', '') 
                                
                                # Montar texto da anotação com quebras de linha
                                texto_anotacao = ""
                                
                                if pd.notna(descricao) and str(descricao).strip() != "":
                                    unidade_texto = f" ({str(unidade).strip()})" if pd.notna(unidade) and str(unidade).strip() != "" else ""
                                    texto_anotacao += f"\n\nDESCRIÇÃO COMPLETA: \n{str(descricao).strip()} {unidade_texto}\n\n"
                                
                                if pd.notna(area_responsavel) and str(area_responsavel).strip() != "":
                                    texto_anotacao += f"Área Responsável: {str(area_responsavel).strip()}\n\n"                                
                                
                                if pd.notna(formula) and str(formula).strip() != "":
                                    texto_anotacao += f"FÓRMULA: {str(formula).strip()}\n\n"
                                
                                if pd.notna(referencia) and str(referencia).strip() != "":
                                    texto_anotacao += f"REFERÊNCIA: \n{str(referencia).strip()}"
                                
                                # Adicionar anotação se houver conteúdo
                                if texto_anotacao.strip():
                                    comment = Comment(texto_anotacao, "Sistema")
                                    
                                    # Calcular altura dinâmica baseada no tamanho do texto
                                    num_linhas = texto_anotacao.count('\n') + 1
                                    altura_dinamica = max(750, num_linhas * 50)  # Mínimo 100, 20 pixels por linha
                                    
                                    # Calcular largura dinâmica baseada no comprimento da linha mais longa
                                    linhas = texto_anotacao.split('\n')
                                    linha_mais_longa = max(len(linha) for linha in linhas) if linhas else 0
                                    largura_dinamica = max(largura_anotacao_min, linha_mais_longa)
                                    
                                    # Aplicar dimensões
                                    comment.width = largura_dinamica
                                    comment.height = altura_dinamica
                                    
                                    # Posicionar a anotação
                                    comment.moveWith = True  # Anotação se move com a célula
                                    
                                    cell2.comment = comment
                        
                    ws.row_dimensions[1].height = 25
                    ws.row_dimensions[2].height = 20
                    ws.freeze_panes = 'C3'
                    
                    # Dados
                    for row_num, (idx, row) in enumerate(df_pivot.iterrows(), 3):
                        for col_num, col_name in enumerate(headers, 1):
                            cell = ws.cell(row=row_num, column=col_num)
                            value = row[col_name]
                            
                            if col_num == 1:
                                cell.value = value
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            elif col_num == 2:
                                cell.value = value
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            else:
                                if pd.notna(value) and str(value).strip() != "":
                                    try:
                                        cell.value = float(value)
                                        cell.number_format = '#,##0.00'
                                    except:
                                        cell.value = value
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            
                            cell.font = data_font
                            cell.border = thin_border
                            
                    ws.column_dimensions['A'].width = 18
                    ws.column_dimensions['B'].width = 35
                    for col_num in range(3, len(headers) + 1):
                        ws.column_dimensions[get_column_letter(col_num)].width = 18
                        
                if len(wb.sheetnames) > 1:
                    wb.remove(ws_primeiro)
                
                modulo_nome = str(modulo).lower().replace(' ', '_')
                tipo_nome = str(info_tipo).lower().replace(' ', '_')
                
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                zip_file.writestr(f'{modulo_nome}_{tipo_nome}.xlsx', buffer.getvalue())
                
    zip_buffer.seek(0)
    return zip_buffer






def criar_tabela_consolidada_html(df_para_tabela, indicadores_filtrados, coluna_valor='anual', incluir_colunas_vazias=False):
    """Cria tabela HTML consolidada com Município e Indicadores"""
    df_pivot_html = df_para_tabela[['codigo_ibge_municipio', 'municipio', 'codigo_da_informacao_indicador', 
                                     'nome_da_informacao_indicador', 'unidade', coluna_valor]].copy()
        
    df_pivot_html = df_pivot_html.drop_duplicates(subset=['codigo_ibge_municipio', 'municipio', 'codigo_da_informacao_indicador'])
    
    # 1. Definir quais indicadores DEVEM aparecer
    if len(indicadores_filtrados) > 0:
        indicadores_esperados = list(indicadores_filtrados)
        df_pivot_html = df_pivot_html[df_pivot_html['codigo_da_informacao_indicador'].isin(indicadores_esperados)]
    else:
        # Se não filtrou, espera todos que estão no DataFrame original
        indicadores_esperados = df_pivot_html['codigo_da_informacao_indicador'].unique().tolist()
        
    # Salvar nomes e unidades
    codigo_nome_unidade_html = {}
    for _, row in df_pivot_html.iterrows():
        codigo = row['codigo_da_informacao_indicador']
        codigo_nome_unidade_html[codigo] = formatar_nome_com_unidade(row['nome_da_informacao_indicador'], row['unidade'])
    
    # 2. Pivot Table (sem dropna=False para evitar erros do Pandas)
    df_pivot_html = df_pivot_html.pivot_table(
        index=['codigo_ibge_municipio', 'municipio'],
        columns='codigo_da_informacao_indicador',
        values=coluna_valor,
        aggfunc='first',
        fill_value=None
    ).reset_index()
    
    # 3. REINSERIR COLUNAS QUE O PIVOT TABLE APAGOU POR ESTAREM VAZIAS
    for indicador in indicadores_esperados:
        if indicador not in df_pivot_html.columns:
            df_pivot_html[indicador] = None
            
    # 4. FILTRAR COLUNAS VAZIAS (SE O CHECKBOX ESTIVER DESMARCADO)
    if not incluir_colunas_vazias:
        colunas_com_dados = ['codigo_ibge_municipio', 'municipio']
        for col in df_pivot_html.columns:
            if col not in ['codigo_ibge_municipio', 'municipio']:
                tem_dado = False
                for val in df_pivot_html[col]:
                    # Verificação robusta contra NaN, None e strings "nan"
                    if pd.notna(val) and val is not None and str(val).strip() != "" and str(val).strip() != "nan" and str(val).strip() != "None":
                        tem_dado = True
                        break
                if tem_dado:
                    colunas_com_dados.append(col)
        df_pivot_html = df_pivot_html[colunas_com_dados]
    else:
        # Se incluir colunas vazias, apenas ordena para ficar na ordem correta
        colunas_ordenadas = ['codigo_ibge_municipio', 'municipio'] + [col for col in indicadores_esperados if col in df_pivot_html.columns]
        df_pivot_html = df_pivot_html[colunas_ordenadas]
    
    indicadores = [col for col in df_pivot_html.columns if col not in ['codigo_ibge_municipio', 'municipio']]
    
    html = '<div style="overflow-x: auto; margin: 0.5rem 0; border-radius: 4px; border: 1px solid #d0d8e0;">'
    html += '<table style="width: 100%; border-collapse: collapse; font-size: 0.8rem;">'
    
    # Cabeçalho linha 1 - Descrição
    html += '<thead><tr style="background-color: #104861; color: white; position: sticky; top: 0; z-index: 2;">'
    html += '<th style="padding: 0.25rem 0.3rem; text-align: center; border: 1px solid #d0d8e0; min-width: 110px; font-weight: 700; font-size: 0.8rem; white-space: normal; position: sticky; left: 0; z-index: 3; background-color: #104861;">Município</th>'
    
    for indicador in indicadores:
        nome = codigo_nome_unidade_html.get(indicador, "")
        html += f'<th style="padding: 0.25rem 0.3rem; text-align: center; border: 1px solid #d0d8e0; min-width: 110px; font-weight: 600; font-size: 0.85rem; white-space: normal;">{nome}</th>'
        
    html += '</tr></thead>'
    
    # Cabeçalho linha 2 - Código
    html += '<thead><tr style="background-color: #0B3040; color: white; position: sticky; top: 28px; z-index: 2;">'
    html += '<th style="padding: 0.25rem 0.3rem; text-align: center; border: 1px solid #d0d8e0; min-width: 110px; font-weight: 700; font-size: 0.8rem; white-space: nowrap; position: sticky; left: 0; z-index: 3; background-color: #0B3040;"></th>'
    
    for indicador in indicadores:
        html += f'<th style="padding: 0.25rem 0.3rem; text-align: center; border: 1px solid #d0d8e0; min-width: 110px; font-weight: 600; font-size: 0.85rem; white-space: nowrap;">{indicador}</th>'
        
    html += '</tr></thead>'
    
    # Corpo da tabela
    html += '<tbody>'
    for idx, row in df_pivot_html.iterrows():
        html += '<tr style="background-color: #ffffff; border-bottom: 1px solid #d0d8e0; height: 16px;">'
        html += f'<td style="padding: 0.2rem 0.3rem; text-align: left; border: 1px solid #d0d8e0; font-weight: 700; font-size: 0.85rem; white-space: nowrap; position: sticky; left: 0; z-index: 1; background-color: #f0f0f0;">{row["municipio"]}</td>'
        
        for indicador in indicadores:
            valor = row[indicador]
            
            # ===== SIMPLES: SE FOR NaN/None, DEIXA VAZIO. SENÃO, FORMATA =====
            if pd.isna(valor) or valor is None or str(valor).strip() == "" or str(valor).strip() == "nan" or str(valor).strip() == "None":
                valor_formatado = ""
            else:
                valor_formatado = formatar_brasileiro(valor)
            
            html += f'<td style="padding: 0.2rem 0.3rem; text-align: right; border: 1px solid #d0d8e0; font-family: \'Calibri\', monospace; font-weight: 500; font-size: 0.85rem; white-space: nowrap;">{valor_formatado}</td>'
            
        html += '</tr>'
        
    html += '</tbody></table></div>'
    return html


# ============================================================================
# FUNÇÕES DE EXPORTAÇÃO ZIP
# ============================================================================

@st.cache_data
def criar_exportacao_zip_por_area(df_sinisa_tuple, coluna_valor='anual', incluir_colunas_vazias=False, largura_anotacao_min=500):
    """Cria ZIP com dados por área responsável - Estilo padronizado com o consolidado
    
    Args:
        df_sinisa_tuple: Dados SINISA
        coluna_valor: Coluna de valor a usar (padrão: 'anual')
        incluir_colunas_vazias: Se True, inclui colunas vazias (padrão: False)
        largura_anotacao_min: Largura mínima da anotação em caracteres (padrão: 500)
    """
    df_sinisa = pd.DataFrame(df_sinisa_tuple)
    zip_buffer = BytesIO()
    
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
    import zipfile
    
    # Carregar glossário para obter Fórmula e Referência
    df_glossario = load_data()
    
    # Estilos baseados no relatório consolidado
    header_fill = PatternFill(start_color='0B3040', end_color='0B3040', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=11)
    thin_border = Border(
        left=Side(style='thin', color='D0D8E0'),
        right=Side(style='thin', color='D0D8E0'),
        top=Side(style='thin', color='D0D8E0'),
        bottom=Side(style='thin', color='D0D8E0')
    )
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # MUDANÇA PRINCIPAL: Agrupar por Área Responsável PRIMEIRO
        areas_responsaveis = sorted(df_sinisa['area_responsavel'].dropna().unique().tolist())
        
        for area in areas_responsaveis:
            df_area = df_sinisa[df_sinisa['area_responsavel'] == area].copy()
            if len(df_area) == 0:
                continue
            
            # Para cada área, criar dois arquivos: um para Água e outro para Esgoto
            for modulo in ['Água', 'Esgoto']:
                df_modulo = df_area[df_area['modulo'] == modulo].copy()
                if len(df_modulo) == 0:
                    continue
                    
                wb = Workbook()
                ws_primeiro = wb.active
                ws_primeiro.title = "Índice"
                
                for grupo in sorted(df_modulo['subformulario_grupo'].dropna().unique().tolist()):
                    df_grupo = df_modulo[df_modulo['subformulario_grupo'] == grupo]
                    codigo_nome_unidade = {}
                    
                    for _, row in df_grupo.iterrows():
                        codigo = row['codigo_da_informacao_indicador']
                        codigo_nome_unidade[codigo] = formatar_nome_com_unidade(row['nome_da_informacao_indicador'], row['unidade'])
                    
                    # OBTER LISTA DE TODOS OS INDICADORES DESTE GRUPO (ANTES DO PIVOT)
                    indicadores_do_grupo = sorted(df_grupo['codigo_da_informacao_indicador'].unique().tolist())
                        
                    df_pivot = df_grupo[['codigo_ibge_municipio', 'municipio', 'codigo_da_informacao_indicador', coluna_valor]].copy()
                    df_pivot = df_pivot.drop_duplicates(subset=['codigo_ibge_municipio', 'municipio', 'codigo_da_informacao_indicador'])
                    df_pivot = df_pivot.pivot_table(
                        index=['codigo_ibge_municipio', 'municipio'],
                        columns='codigo_da_informacao_indicador',
                        values=coluna_valor,
                        aggfunc='first',
                        fill_value=None
                    ).reset_index()
                    
                    # REINSERIR COLUNAS QUE O PIVOT TABLE APAGOU POR ESTAREM VAZIAS
                    for indicador in indicadores_do_grupo:
                        if indicador not in df_pivot.columns:
                            df_pivot[indicador] = None
                    
                    # ===== FILTRAR COLUNAS VAZIAS COM VERIFICAÇÃO ROBUSTA =====
                    if not incluir_colunas_vazias:
                        colunas_com_dados = ['codigo_ibge_municipio', 'municipio']
                        for col in df_pivot.columns:
                            if col not in ['codigo_ibge_municipio', 'municipio']:
                                tem_dado = False
                                for val in df_pivot[col]:
                                    if pd.notna(val) and val is not None and str(val).strip() != "" and str(val).strip() != "nan" and str(val).strip() != "None":
                                        tem_dado = True
                                        break
                                if tem_dado:
                                    colunas_com_dados.append(col)
                        df_pivot = df_pivot[colunas_com_dados]
                    else:
                        colunas_ordenadas = ['codigo_ibge_municipio', 'municipio'] + [col for col in indicadores_do_grupo if col in df_pivot.columns]
                        df_pivot = df_pivot[colunas_ordenadas]
                        
                    ws = wb.create_sheet(title=str(grupo)[:31])
                    headers = list(df_pivot.columns)
                    
                    # Cabeçalhos
                    for col_num, header in enumerate(headers, 1):
                        # Linha 1
                        cell1 = ws.cell(row=1, column=col_num)
                        cell1.value = "" if col_num <= 2 else codigo_nome_unidade.get(header, "")
                        cell1.font = header_font
                        cell1.fill = header_fill
                        cell1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell1.border = thin_border
                        
                        # Linha 2
                        cell2 = ws.cell(row=2, column=col_num)
                        cell2.value = header
                        cell2.font = header_font
                        cell2.fill = header_fill
                        cell2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell2.border = thin_border
                        
                        # ADICIONAR ANOTAÇÃO NA SEGUNDA LINHA
                        if col_num == 1:
                            cell2.comment = Comment("X", "Sistema")
                        elif col_num > 2:
                            codigo_info = header
                            row_glossario = df_glossario[df_glossario['Código da informação'] == codigo_info]
                            
                            if not row_glossario.empty:
                                descricao = row_glossario.iloc[0].get('Descrição SINISA', '')
                                formula = row_glossario.iloc[0].get('Fórmula', '')
                                referencia = row_glossario.iloc[0].get('Referência', '')
                                unidade = row_glossario.iloc[0].get('Unidade', '')
                                
                                # Montar texto da anotação com quebras de linha
                                texto_anotacao = ""
                                
                                if pd.notna(descricao) and str(descricao).strip() != "":
                                    unidade_texto = f" ({str(unidade).strip()})" if pd.notna(unidade) and str(unidade).strip() != "" else ""
                                    texto_anotacao += f"\n\nDESCRIÇÃO COMPLETA: \n{str(descricao).strip()} {unidade_texto}\n\n"
                                
                                if pd.notna(formula) and str(formula).strip() != "":
                                    texto_anotacao += f"FÓRMULA: {str(formula).strip()}\n\n"
                                
                                if pd.notna(referencia) and str(referencia).strip() != "":
                                    texto_anotacao += f"REFERÊNCIA: \n{str(referencia).strip()}"
                                
                                # Adicionar anotação se houver conteúdo
                                if texto_anotacao.strip():
                                    comment = Comment(texto_anotacao, "Sistema")
                                    
                                    # Calcular altura dinâmica baseada no tamanho do texto
                                    num_linhas = texto_anotacao.count('\n') + 1
                                    altura_dinamica = max(750, num_linhas * 50)
                                    
                                    # Calcular largura dinâmica baseada no comprimento da linha mais longa
                                    linhas = texto_anotacao.split('\n')
                                    linha_mais_longa = max(len(linha) for linha in linhas) if linhas else 0
                                    largura_dinamica = max(largura_anotacao_min, linha_mais_longa)
                                    
                                    # Aplicar dimensões
                                    comment.width = largura_dinamica
                                    comment.height = altura_dinamica
                                    comment.moveWith = True
                                    
                                    cell2.comment = comment
                        
                    ws.row_dimensions[1].height = 25
                    ws.row_dimensions[2].height = 20
                    ws.freeze_panes = 'C3'
                    
                    # Dados
                    for row_num, (idx, row) in enumerate(df_pivot.iterrows(), 3):
                        for col_num, col_name in enumerate(headers, 1):
                            cell = ws.cell(row=row_num, column=col_num)
                            value = row[col_name]
                            
                            if col_num == 1:
                                cell.value = value
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            elif col_num == 2:
                                cell.value = value
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            else:
                                if pd.notna(value) and str(value).strip() != "":
                                    try:
                                        cell.value = float(value)
                                        cell.number_format = '#,##0.00'
                                    except:
                                        cell.value = value
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            
                            cell.font = data_font
                            cell.border = thin_border
                            
                    ws.column_dimensions['A'].width = 18
                    ws.column_dimensions['B'].width = 35
                    for col_num in range(3, len(headers) + 1):
                        ws.column_dimensions[get_column_letter(col_num)].width = 18
                        
                if len(wb.sheetnames) > 1:
                    wb.remove(ws_primeiro)
                
                # Salvar arquivo no ZIP com nome: area_modulo.xlsx
                area_nome = str(area).lower().replace(' ', '_').replace('/', '_')
                modulo_nome = str(modulo).lower().replace(' ', '_')
                
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                zip_file.writestr(f'{area_nome}_{modulo_nome}.xlsx', buffer.getvalue())
                
    zip_buffer.seek(0)
    return zip_buffer


def criar_zip_relatorios_todos_municipios(df_agems, ano_selecionado, modulo_selecionado, formulario_selecionado, subformulario_selecionado, subgrupo_selecionado, busca_unificada):
    """Cria ZIP com relatórios Excel para todos os municípios + arquivo Dados_SINISA filtrado"""
    from zipfile import ZipFile
    
    zip_buffer = BytesIO()
    
    # Aplicar filtros UMA VEZ para todo o dataframe
    df_filtrado = aplicar_filtros(df_agems, {
        'modulo': modulo_selecionado,
        'formulario': formulario_selecionado,
        'subformulario_grupo': subformulario_selecionado,
        'subgrupo_palavra_chave': subgrupo_selecionado
    })
    
    # Aplicar busca unificada UMA VEZ
    if busca_unificada:
        df_filtrado = df_filtrado[
            (df_filtrado['codigo_da_informacao_indicador'].str.contains(busca_unificada, case=False, na=False)) |
            (df_filtrado['nome_da_informacao_indicador'].str.contains(busca_unificada, case=False, na=False))
        ]
    
    # Obter colunas de meses UMA VEZ
    todas_colunas_meses = sorted([col for col in df_filtrado.columns if '/' in col and len(col) == 7])
    colunas_meses = [mes for mes in todas_colunas_meses if len(df_filtrado[mes].dropna()) > 0]
    
    if len(colunas_meses) == 0:
        return zip_buffer
    
    # Obter lista de municípios que têm dados após filtros
    municipios = sorted(df_filtrado['municipio'].dropna().unique().tolist())
    
    with ZipFile(zip_buffer, 'w', compression=8) as zip_file:
        # ===== GERAR RELATÓRIOS EXCEL PARA CADA MUNICÍPIO =====
        for municipio in municipios:
            try:
                df_municipio = df_filtrado[df_filtrado['municipio'] == municipio].copy()
                
                if len(df_municipio) > 0:
                    wb = criar_relatorio_excel_generico(df_municipio, municipio, colunas_meses)
                    buffer = BytesIO()
                    wb.save(buffer)
                    buffer.seek(0)
                    
                    nome_arquivo = f"RelGer {municipio.replace(' ', '_')}_{ano_selecionado}.xlsx"
                    zip_file.writestr(nome_arquivo, buffer.getvalue())
            except Exception as e:
                continue
        
        # =====         # ===== ADICIONAR ARQUIVO DADOS_SINISA FILTRADO (OTIMIZADO) =====
        try:
            file_path = os.path.join(DATA_DIR, "Dados_SINISA.xlsx")
            
            # Ler a aba correspondente ao ano selecionado
            df_dados_sinisa = pd.read_excel(file_path, sheet_name=str(ano_selecionado))
            df_dados_sinisa.columns = df_dados_sinisa.columns.str.strip()
            
            # Colunas a remover
            colunas_remover = [
                'Área Responsável',
                'Acumulado em 12 meses',
                'Média em 12 meses',
                'Relatório Gerencial - AGEMS'
            ]
            
            # Remover colunas que existem
            colunas_existentes = [col for col in colunas_remover if col in df_dados_sinisa.columns]
            df_dados_sinisa = df_dados_sinisa.drop(columns=colunas_existentes)
            
            # OTIMIZAÇÃO: Usar pandas.to_excel em vez de iterar célula por célula
            buffer_dados = BytesIO()
            
            # Usar o ExcelWriter do pandas (muito mais rápido)
            with pd.ExcelWriter(buffer_dados, engine='openpyxl') as writer:
                df_dados_sinisa.to_excel(writer, sheet_name=f"Dados_{ano_selecionado}", index=False)
                
                # Aplicar estilo apenas ao cabeçalho para manter a performance alta
                workbook = writer.book
                worksheet = writer.sheets[f"Dados_{ano_selecionado}"]
                
                from openpyxl.styles import PatternFill, Font
                from openpyxl.utils import get_column_letter
                
                header_fill = PatternFill(start_color='0B3040', end_color='0B3040', fill_type='solid')
                header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                
                # Estilizar apenas a primeira linha (cabeçalho)
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    
                # Ajustar largura das colunas de forma rápida
                for col_num in range(1, len(df_dados_sinisa.columns) + 1):
                    worksheet.column_dimensions[get_column_letter(col_num)].width = 20

            buffer_dados.seek(0)
            
            # Adicionar ao ZIP
            nome_arquivo_dados = f"Dados_SINISA_{ano_selecionado}.xlsx"
            zip_file.writestr(nome_arquivo_dados, buffer_dados.getvalue())
            
        except Exception as e:
            # Se houver erro ao adicionar Dados_SINISA, continua mesmo assim
            pass
    
    zip_buffer.seek(0)
    return zip_buffer

# ============================================================================
# FUNÇÕES DE PÁGINAS
# ============================================================================
# Cache global para armazenar ZIP gerado
_zip_cache = {}

@st.cache_data(ttl=3600)
def gerar_zip_background(df_agems_tuple, ano_selecionado, modulo_selecionado, formulario_selecionado, subformulario_selecionado, subgrupo_selecionado, busca_unificada):
    """Gera ZIP em background e armazena em cache"""
    # Converter tuple de volta para DataFrame
    df_agems = pd.DataFrame(df_agems_tuple)
    
    zip_buffer = BytesIO()
    
    # Aplicar filtros UMA VEZ para todo o dataframe
    df_filtrado = aplicar_filtros(df_agems, {
        'modulo': modulo_selecionado,
        'formulario': formulario_selecionado,
        'subformulario_grupo': subformulario_selecionado,
        'subgrupo_palavra_chave': subgrupo_selecionado
    })
    
    # Aplicar busca unificada UMA VEZ
    if busca_unificada:
        df_filtrado = df_filtrado[
            (df_filtrado['codigo_da_informacao_indicador'].str.contains(busca_unificada, case=False, na=False)) |
            (df_filtrado['nome_da_informacao_indicador'].str.contains(busca_unificada, case=False, na=False))
        ]
    
    # Obter colunas de meses UMA VEZ
    todas_colunas_meses = sorted([col for col in df_filtrado.columns if '/' in col and len(col) == 7])
    colunas_meses = [mes for mes in todas_colunas_meses if len(df_filtrado[mes].dropna()) > 0]
    
    if len(colunas_meses) == 0:
        return zip_buffer
    
    # Obter lista de municípios que têm dados após filtros
    municipios = sorted(df_filtrado['municipio'].dropna().unique().tolist())
    
    from zipfile import ZipFile
    with ZipFile(zip_buffer, 'w', compression=8) as zip_file:
        for municipio in municipios:
            try:
                df_municipio = df_filtrado[df_filtrado['municipio'] == municipio].copy()
                
                if len(df_municipio) > 0:
                    wb = criar_relatorio_excel_generico(df_municipio, municipio, colunas_meses)
                    buffer = BytesIO()
                    wb.save(buffer)
                    buffer.seek(0)
                    
                    nome_arquivo = f"relatorio_{municipio.replace(' ', '_')}_{ano_selecionado}.xlsx"
                    zip_file.writestr(nome_arquivo, buffer.getvalue())
            except Exception as e:
                continue
    
    zip_buffer.seek(0)
    return zip_buffer
# ============================================================================
# FUNÇÃO DE EXIBIÇÃO DE DETALHES EM POPUP MODAL (st.dialog)
# ============================================================================


@st.dialog("📋 Detalhes", width="large")
def exibir_detalhes_popup(row):
    """Exibe detalhes em popup modal com verificações robustas de campos vazios"""
    
    # Função auxiliar para verificar se o valor é realmente válido e não vazio
    def get_valor_valido(valor):
        if pd.isna(valor):
            return ""
        texto = str(valor).strip()
        if texto.lower() in ['nan', 'none', '']:
            return ""
        return texto

    # Título compacto
    st.markdown(f"""
    <div style="background: #d4e4f7; border-radius: 6px; padding: 0.8rem; margin-bottom: 0.8rem; border-left: 4px solid #1f4788;">
        <h4 style="color: #1f4788; margin: 0; font-size: 1.1rem;">
            <strong>{row['Código da informação']}</strong> — {row['Nome da informação']}
        </h4>
    </div>
    """, unsafe_allow_html=True)
    
    # Linha única com 5 colunas
    col1, col2, col3, col4, col5 = st.columns(5, gap="small")
    
    with col1:
        st.markdown(f"""
        <div style="font-size: 0.8rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Módulo</div>
        <div style="font-size: 0.95rem; color: #333; border-bottom: 1px solid #c0d9f0; padding-bottom: 0.3rem; line-height: 1.3;">{get_valor_valido(row['Módulo']) or 'N/A'}</div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
        <div style="font-size: 0.8rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Formulário</div>
        <div style="font-size: 0.95rem; color: #333; border-bottom: 1px solid #c0d9f0; padding-bottom: 0.3rem; line-height: 1.3;">{get_valor_valido(row['Formulário']) or 'N/A'}</div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown(f"""
        <div style="font-size: 0.8rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Subformulário</div>
        <div style="font-size: 0.95rem; color: #333; border-bottom: 1px solid #c0d9f0; padding-bottom: 0.3rem; line-height: 1.3;">{get_valor_valido(row['Subformulário']) or 'N/A'}</div>
        """, unsafe_allow_html=True)
    
    with col4:
        st.markdown(f"""
        <div style="font-size: 0.8rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Subgrupo</div>
        <div style="font-size: 0.95rem; color: #333; border-bottom: 1px solid #c0d9f0; padding-bottom: 0.3rem; line-height: 1.3;">{get_valor_valido(row['Subgrupo']) or 'N/A'}</div>
        """, unsafe_allow_html=True)
    
    with col5:
        st.markdown(f"""
        <div style="font-size: 0.8rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Área Responsável</div>
        <div style="font-size: 0.95rem; color: #333; border-bottom: 1px solid #c0d9f0; padding-bottom: 0.3rem; line-height: 1.3;">{get_valor_valido(row['Área Responsável']) or 'N/A'}</div>
        """, unsafe_allow_html=True)
    

    
    # Descrição e Unidade
    val_descricao = get_valor_valido(row['Descrição SINISA'])
    descricao_texto = val_descricao if val_descricao else 'Não informada'
    
    val_unidade = get_valor_valido(row['Unidade'])
    unidade_texto = f"<br><span style='color: #1f4788;'><strong>[Unidade: {val_unidade}]</strong></span>" if val_unidade else ""
    
    st.markdown(f"""
    <div>
        <div style="font-size: 0.85rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Descrição</div>
        <div style="font-size: 1rem; color: #333; border-bottom: 1px solid #c0d9f0; padding-bottom: 0.4rem; line-height: 1.5;">
            {descricao_texto}{unidade_texto}
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Fórmula
    val_formula = get_valor_valido(row['Fórmula'])
    formula_texto = val_formula if val_formula else 'Não informada'
    st.markdown(f"""
    <div style="margin-top: 0.8rem;">
        <div style="font-size: 0.85rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Fórmula</div>
        <div style="background: #c0d9f0; padding: 0.6rem; border-radius: 4px; font-family: monospace; font-size: 0.95rem; color: #333; line-height: 1.4; overflow-x: auto;">
            {formula_texto}
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Referência
    val_referencia = get_valor_valido(row['Referência'])
    if val_referencia:
        if '\n' in val_referencia:
            referencia_texto = val_referencia.replace('\n', '<br>')
        elif ', ' in val_referencia:
            refs = [r.strip() for r in val_referencia.split(',')]
            referencia_texto = "<br>".join([f"• {r}" for r in refs if r])
        else:
            referencia_texto = val_referencia
    else:
        referencia_texto = 'Não informada'
        
    st.markdown(f"""
    <div style="margin-top: 0.8rem;">
        <div style="font-size: 0.85rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Informações/Referência</div>
        <div style="background: #c0d9f0; padding: 0.6rem; border-radius: 4px; font-size: 0.95rem; color: #333; line-height: 1.6;">
            {referencia_texto}
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Fonte
    val_fonte = get_valor_valido(row['Fonte'])
    fonte_texto = val_fonte if val_fonte else 'Não informada'
    st.markdown(f"""
    <div style="margin-top: 0.8rem;">
        <div style="font-size: 0.85rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Fonte</div>
        <div style="font-size: 1rem; color: #333; border-bottom: 1px solid #c0d9f0; padding-bottom: 0.4rem;">
            {fonte_texto}
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Observação (Oculta se vazia e com tratamento para evitar erro de HTML)
    val_observacao = get_valor_valido(row['Observação'])
    if val_observacao:
        obs_texto = val_observacao.replace('\n', '<br>')
        st.markdown(f"""
        <div style="margin-top: 0.8rem;">
            <div style="font-size: 0.85rem; color: #c41e3a; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">⚠️ Observação</div>
            <div style="background: #ffe6e6; padding: 0.6rem; border-radius: 4px; border-left: 3px solid #c41e3a; font-size: 0.95rem; color: #333; line-height: 1.5;">{obs_texto}</div>
        </div>
        """, unsafe_allow_html=True)


def exibir_detalhes_modal(row, idx):
    """Exibe detalhes em popup modal com verificações robustas de campos vazios"""
    
    # Função auxiliar para verificar se o valor é realmente válido e não vazio
    def get_valor_valido(valor):
        if pd.isna(valor):
            return ""
        texto = str(valor).strip()
        if texto.lower() in ['nan', 'none', '']:
            return ""
        return texto

    # Título compacto
    st.markdown(f"""
    <div style="background: #d4e4f7; border-radius: 6px; padding: 0.8rem; margin-bottom: 0.8rem; border-left: 4px solid #1f4788;">
        <h4 style="color: #1f4788; margin: 0; font-size: 1.1rem;">
            <strong>{row['Código da informação']}</strong> — {row['Nome da informação']}
        </h4>
    </div>
    """, unsafe_allow_html=True)
    
    # Linha única com 5 colunas
    col1, col2, col3, col4, col5 = st.columns(5, gap="small")
    
    with col1:
        st.markdown(f"""
        <div style="font-size: 0.8rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Módulo</div>
        <div style="font-size: 0.95rem; color: #333; border-bottom: 1px solid #c0d9f0; padding-bottom: 0.3rem; line-height: 1.3;">{get_valor_valido(row['Módulo']) or 'N/A'}</div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
        <div style="font-size: 0.8rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Formulário</div>
        <div style="font-size: 0.95rem; color: #333; border-bottom: 1px solid #c0d9f0; padding-bottom: 0.3rem; line-height: 1.3;">{get_valor_valido(row['Formulário']) or 'N/A'}</div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown(f"""
        <div style="font-size: 0.8rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Subformulário</div>
        <div style="font-size: 0.95rem; color: #333; border-bottom: 1px solid #c0d9f0; padding-bottom: 0.3rem; line-height: 1.3;">{get_valor_valido(row['Subformulário']) or 'N/A'}</div>
        """, unsafe_allow_html=True)
    
    with col4:
        st.markdown(f"""
        <div style="font-size: 0.8rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Subgrupo</div>
        <div style="font-size: 0.95rem; color: #333; border-bottom: 1px solid #c0d9f0; padding-bottom: 0.3rem; line-height: 1.3;">{get_valor_valido(row['Subgrupo']) or 'N/A'}</div>
        """, unsafe_allow_html=True)
    
    with col5:
        st.markdown(f"""
        <div style="font-size: 0.8rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Área Responsável</div>
        <div style="font-size: 0.95rem; color: #333; border-bottom: 1px solid #c0d9f0; padding-bottom: 0.3rem; line-height: 1.3;">{get_valor_valido(row['Área Responsável']) or 'N/A'}</div>
        """, unsafe_allow_html=True)
    

    
    # Descrição e Unidade
    val_descricao = get_valor_valido(row['Descrição SINISA'])
    descricao_texto = val_descricao if val_descricao else 'Não informada'
    
    val_unidade = get_valor_valido(row['Unidade'])
    unidade_texto = f"<br><span style='color: #1f4788;'><strong>[Unidade: {val_unidade}]</strong></span>" if val_unidade else ""
    
    st.markdown(f"""
    <div>
        <div style="font-size: 0.85rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Descrição</div>
        <div style="font-size: 1rem; color: #333; border-bottom: 1px solid #c0d9f0; padding-bottom: 0.4rem; line-height: 1.5;">
            {descricao_texto}{unidade_texto}
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Fórmula
    val_formula = get_valor_valido(row['Fórmula'])
    formula_texto = val_formula if val_formula else 'Não informada'
    st.markdown(f"""
    <div style="margin-top: 0.8rem;">
        <div style="font-size: 0.85rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Fórmula</div>
        <div style="background: #c0d9f0; padding: 0.6rem; border-radius: 4px; font-family: monospace; font-size: 0.95rem; color: #333; line-height: 1.4; overflow-x: auto;">
            {formula_texto}
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Referência
    val_referencia = get_valor_valido(row['Referência'])
    if val_referencia:
        if '\n' in val_referencia:
            referencia_texto = val_referencia.replace('\n', '<br>')
        elif ', ' in val_referencia:
            refs = [r.strip() for r in val_referencia.split(',')]
            referencia_texto = "<br>".join([f"• {r}" for r in refs if r])
        else:
            referencia_texto = val_referencia
    else:
        referencia_texto = 'Não informada'
        
    st.markdown(f"""
    <div style="margin-top: 0.8rem;">
        <div style="font-size: 0.85rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Informações/Referência</div>
        <div style="background: #c0d9f0; padding: 0.6rem; border-radius: 4px; font-size: 0.95rem; color: #333; line-height: 1.6;">
            {referencia_texto}
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Fonte
    val_fonte = get_valor_valido(row['Fonte'])
    fonte_texto = val_fonte if val_fonte else 'Não informada'
    st.markdown(f"""
    <div style="margin-top: 0.8rem;">
        <div style="font-size: 0.85rem; color: #1f4788; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">Fonte</div>
        <div style="font-size: 1rem; color: #333; border-bottom: 1px solid #c0d9f0; padding-bottom: 0.4rem;">
            {fonte_texto}
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Observação (Oculta se vazia e com tratamento para evitar erro de HTML)
    val_observacao = get_valor_valido(row['Observação'])
    if val_observacao:
        obs_texto = val_observacao.replace('\n', '<br>')
        st.markdown(f"""
        <div style="margin-top: 0.8rem;">
            <div style="font-size: 0.85rem; color: #c41e3a; font-weight: 700; text-transform: uppercase; margin-bottom: 0.2rem; letter-spacing: 0.5px;">⚠️ Observação</div>
            <div style="background: #ffe6e6; padding: 0.6rem; border-radius: 4px; border-left: 3px solid #c41e3a; font-size: 0.95rem; color: #333; line-height: 1.5;">{obs_texto}</div>
        </div>
        """, unsafe_allow_html=True)

# ============================================================================
# FUNÇÃO NOVA - FORMATAR EXCEL DO GLOSSÁRIO COM ESTÉTICA DASHBOARD (DESCRIÇÃO SINISA POR ÚLTIMO)
# ============================================================================
def criar_excel_glossario_formatado(df_filtrado):
    """
    Cria Excel formatado para o Glossário de Informações com estética do dashboard
    - Cabeçalhos com cores do dashboard
    - Linhas alternadas
    - Bordas e espaçamento otimizados
    - Colunas: Módulo, Formulário, Subformulário, Subgrupo, Área Responsável, 
      Código da informação, Nome da informação, Unidade, Descrição SINISA, 
      Fonte, Fórmula, Referência, Operação
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossário"
    
    # Definir largura das colunas - NOVA ORDEM
    ws.column_dimensions['A'].width = 9  # Módulo
    ws.column_dimensions['B'].width = 12 # Formulário
    ws.column_dimensions['C'].width = 20  # Subformulário
    ws.column_dimensions['D'].width = 15  # Subgrupo
    ws.column_dimensions['E'].width = 18  # Área Responsável
    ws.column_dimensions['F'].width = 14  # Código da informação
    ws.column_dimensions['G'].width = 30  # Nome da informação
    ws.column_dimensions['H'].width = 12  # Unidade
    ws.column_dimensions['I'].width = 50  # Descrição SINISA
    ws.column_dimensions['J'].width = 18  # Fonte
    ws.column_dimensions['K'].width = 20  # Fórmula
    ws.column_dimensions['L'].width = 18  # Referência
    ws.column_dimensions['M'].width = 14  # Operação
    
    # Definir altura do cabeçalho
    ws.row_dimensions[1].height = 20
    
    # Estilos
    header_fill = PatternFill(
        start_color='0B3040',
        end_color='0B3040',
        fill_type='solid'
    )
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    
    row_alternado_fill = PatternFill(
        start_color='F5F7FA',
        end_color='F5F7FA',
        fill_type='solid'
    )
    row_normal_fill = PatternFill(
        start_color='FFFFFF',
        end_color='FFFFFF',
        fill_type='solid'
    )
    
    border = Border(
        left=Side(style='thin', color='D0D8E0'),
        right=Side(style='thin', color='D0D8E0'),
        top=Side(style='thin', color='D0D8E0'),
        bottom=Side(style='thin', color='D0D8E0')
    )
    
    data_font = Font(name='Calibri', size=10)
    
    # Cabeçalhos - NOVA ORDEM
    headers = ['Módulo', 'Formulário', 'Subformulário', 'Subgrupo', 'Área Responsável', 'Código da informação', 'Nome da informação', 'Unidade', 'Descrição SINISA', 'Fonte', 'Fórmula', 'Referência', 'Operação']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Dados com linhas alternadas - NOVA ORDEM
    alternado = False
    for row_num, (idx, row) in enumerate(df_filtrado.iterrows(), 2):
        # Alternar cores
        bg_fill = row_alternado_fill if alternado else row_normal_fill
        alternado = not alternado
        
        # Coluna A: Módulo
        cell = ws.cell(row=row_num, column=1)
        cell.value = row['Módulo'] if pd.notna(row['Módulo']) else '-'
        cell.font = data_font
        cell.fill = bg_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
        
        # Coluna B: Formulário
        cell = ws.cell(row=row_num, column=2)
        cell.value = row['Formulário'] if pd.notna(row['Formulário']) else '-'
        cell.font = data_font
        cell.fill = bg_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
        
        # Coluna C: Subformulário
        cell = ws.cell(row=row_num, column=3)
        cell.value = row['Subformulário'] if pd.notna(row['Subformulário']) else '-'
        cell.font = data_font
        cell.fill = bg_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
        
        # Coluna D: Subgrupo
        cell = ws.cell(row=row_num, column=4)
        cell.value = row['Subgrupo'] if pd.notna(row['Subgrupo']) else '-'
        cell.font = data_font
        cell.fill = bg_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
        
        # Coluna E: Área Responsável
        cell = ws.cell(row=row_num, column=5)
        cell.value = row['Área Responsável'] if pd.notna(row['Área Responsável']) else '-'
        cell.font = data_font
        cell.fill = bg_fill
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border
        
        # Coluna F: Código da informação
        cell = ws.cell(row=row_num, column=6)
        cell.value = row['Código da informação'] if pd.notna(row['Código da informação']) else '-'
        cell.font = data_font
        cell.fill = bg_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
        
        # Coluna G: Nome da informação
        cell = ws.cell(row=row_num, column=7)
        cell.value = row['Nome da informação'] if pd.notna(row['Nome da informação']) else '-'
        cell.font = data_font
        cell.fill = bg_fill
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border
        
        # Coluna H: Unidade
        cell = ws.cell(row=row_num, column=8)
        cell.value = row['Unidade'] if pd.notna(row['Unidade']) else '-'
        cell.font = data_font
        cell.fill = bg_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        
        # Coluna I: Descrição SINISA
        cell = ws.cell(row=row_num, column=9)
        cell.value = row['Descrição SINISA'] if pd.notna(row['Descrição SINISA']) else '-'
        cell.font = data_font
        cell.fill = bg_fill
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border
        
        # Coluna J: Fonte
        cell = ws.cell(row=row_num, column=10)
        cell.value = row['Fonte'] if pd.notna(row['Fonte']) else '-'
        cell.font = data_font
        cell.fill = bg_fill
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border
        
        # Coluna K: Fórmula
        cell = ws.cell(row=row_num, column=11)
        cell.value = row['Fórmula'] if pd.notna(row['Fórmula']) else '-'
        cell.font = data_font
        cell.fill = bg_fill
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border
        
        # Coluna L: Referência
        cell = ws.cell(row=row_num, column=12)
        cell.value = row['Referência'] if pd.notna(row['Referência']) else '-'
        cell.font = data_font
        cell.fill = bg_fill
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border
        
        # Coluna M: Operação
        cell = ws.cell(row=row_num, column=13)
        cell.value = row['Operação'] if pd.notna(row['Operação']) else '-'
        cell.font = data_font
        cell.fill = bg_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
        
        ws.row_dimensions[row_num].height = 25
    
    return wb

def pagina_sobre():
    """Página Sobre"""
    st.markdown("""
    <div class="header-container">
        <h1 class="header-title">📚 SINISA</h1>
        <p class="header-subtitle">Sistema Nacional de Informações sobre Saneamento</p>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("### 📖 Sobre o Sistema")
    
    st.info("""
    ⏳ **Esta seção está em construção.**
    
    Em breve, você poderá acessar informações detalhadas sobre o SINISA, sua estrutura, objetivos e funcionalidades.
    """)
    
    st.markdown("---")
    
    st.markdown("""
    ### 🔗 Links Úteis
    - [Portal SINISA](https://www.gov.br/cidades/pt-br/acesso-a-informacao/acoes-e-programas/saneamento/sinisa)
    - [Documentação](https://www.gov.br/cidades/pt-br/acesso-a-informacao/acoes-e-programas/saneamento/sinisa/sinisa-1)
    """)

def pagina_glossario_informacoes_com_stats():
    """Página Glossário - Informações com Estatísticas - VERSÃO REFATORADA"""
    st.markdown("""
    <div class="header-container">
        <h1 class="header-title">📚 SINISA</h1>
        <p class="header-subtitle">Sistema Nacional de Informações sobre Saneamento</p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("### 📋 Glossário - Informações")
    
    # Inicializar session state
    if "modulo_filter" not in st.session_state:
        st.session_state.modulo_filter = "Todos"
    if "formulario_filter" not in st.session_state:
        st.session_state.formulario_filter = "Todos"
    if "subformulario_filter" not in st.session_state:
        st.session_state.subformulario_filter = "Todos"
    if "subgrupo_filter" not in st.session_state:
        st.session_state.subgrupo_filter = "Todos"
    if "area_filter" not in st.session_state:
        st.session_state.area_filter = "Todos"
    # NOVO: Session state para campos de busca refatorados
    if "search_codigo_informacao" not in st.session_state:
        st.session_state.search_codigo_informacao = ""
    if "search_descricao_sinisa" not in st.session_state:
        st.session_state.search_descricao_sinisa = ""
    if "show_stats" not in st.session_state:
        st.session_state.show_stats = False
    
    # Funções de filtro
    @st.cache_data
    def get_modulos():
        modulos = ["Todos"] + sort_modulos(df['Módulo'].dropna().unique().tolist())
        return modulos
    
    def get_filtro_glossario(coluna, modulo="Todos", formulario="Todos", subformulario="Todos", subgrupo="Todos"):
        """Obtém valores únicos para filtros do glossário"""
        df_temp = df.copy()
        if modulo != "Todos":
            df_temp = df_temp[df_temp['Módulo'] == modulo]
        if formulario != "Todos":
            df_temp = df_temp[df_temp['Formulário'] == formulario]
        if subformulario != "Todos":
            df_temp = df_temp[df_temp['Subformulário'] == subformulario]
        if subgrupo != "Todos":
            df_temp = df_temp[df_temp['Subgrupo'] == subgrupo]
        return ["Todos"] + sorted(df_temp[coluna].dropna().unique().tolist())
    
    # Linha 1 - Filtros
    col1, col2, col3, col4, col5, col6 = st.columns([1.5, 1.5, 1.5, 1.5, 1.5, 1], gap="small")
    with col6:
        st.markdown("<div style='padding-top: 12px;'></div>", unsafe_allow_html=True)
        if st.button("🔄 Limpar Filtros", key="btn_clear_filters", use_container_width=True):
            st.session_state.modulo_filter = "Todos"
            st.session_state.formulario_filter = "Todos"
            st.session_state.subformulario_filter = "Todos"
            st.session_state.subgrupo_filter = "Todos"
            st.session_state.area_filter = "Todos"
            st.session_state.search_codigo_informacao = ""  # NOVO
            st.session_state.search_descricao_sinisa = ""   # NOVO
            st.rerun()
    
    with col1:
        modulo_selecionado = st.selectbox("Módulo", get_modulos(), key="modulo_filter", index=0)
    with col2:
        formulario_selecionado = st.selectbox("Formulário", 
            get_filtro_glossario('Formulário', modulo_selecionado), key="formulario_filter", index=0)
    with col3:
        subformulario_selecionado = st.selectbox("Subformulário", 
            get_filtro_glossario('Subformulário', modulo_selecionado, formulario_selecionado), 
            key="subformulario_filter", index=0)
    with col4:
        subgrupo_selecionado = st.selectbox("Subgrupo", 
            get_filtro_glossario('Subgrupo', modulo_selecionado, formulario_selecionado, subformulario_selecionado), 
            key="subgrupo_filter", index=0)
    with col5:
        area_selecionada = st.selectbox("Área Responsável", 
            get_filtro_glossario('Área Responsável', modulo_selecionado, formulario_selecionado, 
                                subformulario_selecionado, subgrupo_selecionado), 
            key="area_filter", index=0)
    
    # ========================================================================
    # LINHA 2 - BUSCAS REFATORADAS (ALTERAÇÃO PRINCIPAL)
    # ========================================================================
    col_codigo_info, col_descricao_sinisa = st.columns([2.1, 3.9], gap="small")
    
    with col_codigo_info:
        busca_codigo_informacao = st.text_input(
            "🔎 Buscar por Código/Informação",
            key="search_codigo_informacao",
            placeholder="Digite código ou nome da informação"
        )
    
    with col_descricao_sinisa:
        busca_descricao_sinisa = st.text_input(
            "🔎 Buscar por Palavras-Chaves (Pesquisa Ampla)",
            key="search_descricao_sinisa",
            placeholder="Digite a descrição SINISA"
        )
    
    # Aplicar filtros
    df_filtrado = aplicar_filtros(df, {
        'Módulo': modulo_selecionado,
        'Formulário': formulario_selecionado,
        'Subformulário': subformulario_selecionado,
        'Subgrupo': subgrupo_selecionado,
        'Área Responsável': area_selecionada
    })
    
    # ========================================================================
    # LÓGICA DE BUSCA REFATORADA - ALTERAÇÃO 1: Código/Informação Unificada
    # ========================================================================
    if busca_codigo_informacao.strip():  # Tratamento de entrada vazia
        try:
            # Busca case-insensitive em CÓDIGO ou NOME DA INFORMAÇÃO
            termo_busca = busca_codigo_informacao.strip().lower()
            df_filtrado = df_filtrado[
                (df_filtrado['Código da informação'].astype(str).str.lower().str.contains(termo_busca, na=False)) |
                (df_filtrado['Nome da informação'].astype(str).str.lower().str.contains(termo_busca, na=False))
            ]
        except Exception as e:
            st.warning(f"⚠️ Erro ao processar busca de Código/Informação: {str(e)}")
    
    # ========================================================================
    # LÓGICA DE BUSCA REFATORADA - ALTERAÇÃO 2: Descrição SINISA Específica
    # ========================================================================
    if busca_descricao_sinisa.strip():  # Tratamento de entrada vazia
        try:
            # Busca case-insensitive especificamente na coluna "Descrição SINISA"
            termo_busca = busca_descricao_sinisa.strip().lower()
            df_filtrado = df_filtrado[
                df_filtrado['Descrição SINISA'].astype(str).str.lower().str.contains(termo_busca, na=False)
            ]
        except Exception as e:
            st.warning(f"⚠️ Erro ao processar busca de Descrição SINISA: {str(e)}")
    
    # Função de exibição de estatísticas modal
    def exibir_estatisticas_modal(df_filtrado):
        """Exibe estatísticas dos dados filtrados"""
        st.markdown(f"""
        <div class="detail-section">
            <h3 style="color: #1f4788; margin-top: 0; margin-bottom: 1rem;">📊 Análise dos Dados</h3>
        </div>
        """, unsafe_allow_html=True)
        col1, col2, col3, col4 = st.columns(4, gap="small")
        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">📋 Total</div>
                <div class="metric-value">{len(df_filtrado)}</div>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">📦 Módulos</div>
                <div class="metric-value">{df_filtrado['Módulo'].nunique()}</div>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">📝 Formulários</div>
                <div class="metric-value">{df_filtrado['Formulário'].nunique()}</div>
            </div>
            """, unsafe_allow_html=True)
        with col4:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">🏷️ Subgrupos</div>
                <div class="metric-value">{df_filtrado['Subgrupo'].nunique() if 'Subgrupo' in df_filtrado.columns else 0}</div>
            </div>
            """, unsafe_allow_html=True)
        
        # Gráficos
        col1, col2, col3, col4 = st.columns(4, gap="small")
        with col1:
            if len(df_filtrado) > 0:
                modulo_count = df_filtrado['Módulo'].value_counts().head(6)
                fig = go.Figure(data=[go.Bar(x=modulo_count.values, y=modulo_count.index, orientation='h', 
                    marker=dict(color=modulo_count.values, colorscale='Blues', line=dict(color='#1f4788', width=1)), 
                    text=modulo_count.values, textposition='auto', hovertemplate='<b>%{y}</b><br>%{x}<extra></extra>')])
                fig.update_layout(height=250, margin=dict(l=100, r=10, t=10, b=10), plot_bgcolor='rgba(245, 247, 250, 0.5)', 
                    paper_bgcolor='rgba(255, 255, 255, 0)', xaxis=dict(showgrid=True, gridwidth=1, gridcolor='#e0e0e0', showticklabels=False), 
                    yaxis=dict(showgrid=False), font=dict(family='Segoe UI', size=9, color='#333'), hovermode='closest', showlegend=False)
                st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
        with col2:
            if len(df_filtrado) > 0:
                formulario_count = df_filtrado['Formulário'].value_counts().head(6)
                fig = go.Figure(data=[go.Bar(x=formulario_count.values, y=formulario_count.index, orientation='h', 
                    marker=dict(color=formulario_count.values, colorscale='Greens', line=dict(color='#2d5aa8', width=1)), 
                    text=formulario_count.values, textposition='auto', hovertemplate='<b>%{y}</b><br>%{x}<extra></extra>')])
                fig.update_layout(height=250, margin=dict(l=100, r=10, t=10, b=10), plot_bgcolor='rgba(245, 247, 250, 0.5)', 
                    paper_bgcolor='rgba(255, 255, 255, 0)', xaxis=dict(showgrid=True, gridwidth=1, gridcolor='#e0e0e0', showticklabels=False), 
                    yaxis=dict(showgrid=False), font=dict(family='Segoe UI', size=9, color='#333'), hovermode='closest', showlegend=False)
                st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
        with col3:
            if len(df_filtrado) > 0 and 'Subgrupo' in df_filtrado.columns:
                subgrupo_count = df_filtrado['Subgrupo'].value_counts().head(6)
                fig = go.Figure(data=[go.Bar(x=subgrupo_count.values, y=subgrupo_count.index, orientation='h', 
                    marker=dict(color=subgrupo_count.values, colorscale='Purples', line=dict(color='#6b21a8', width=1)), 
                    text=subgrupo_count.values, textposition='auto', hovertemplate='<b>%{y}</b><br>%{x}<extra></extra>')])
                fig.update_layout(height=250, margin=dict(l=100, r=10, t=10, b=10), plot_bgcolor='rgba(245, 247, 250, 0.5)', 
                    paper_bgcolor='rgba(255, 255, 255, 0)', xaxis=dict(showgrid=True, gridwidth=1, gridcolor='#e0e0e0', showticklabels=False),
                    yaxis=dict(showgrid=False), font=dict(family='Segoe UI', size=9, color='#333'), hovermode='closest', showlegend=False)
                st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
        with col4:
            if len(df_filtrado) > 0 and 'Área Responsável' in df_filtrado.columns:
                area_count = df_filtrado['Área Responsável'].value_counts().head(6)
                fig = go.Figure(data=[go.Bar(x=area_count.values, y=area_count.index, orientation='h', 
                    marker=dict(color=area_count.values, colorscale='Oranges', line=dict(color='#d97706', width=1)), 
                    text=area_count.values, textposition='auto', hovertemplate='<b>%{y}</b><br>%{x}<extra></extra>')])
                fig.update_layout(height=250, margin=dict(l=100, r=10, t=10, b=10), plot_bgcolor='rgba(245, 247, 250, 0.5)',
                    paper_bgcolor='rgba(255, 255, 255, 0)', xaxis=dict(showgrid=True, gridwidth=1, gridcolor='#e0e0e0', showticklabels=False), 
                    yaxis=dict(showgrid=False), font=dict(family='Segoe UI', size=9, color='#333'), hovermode='closest', showlegend=False)
                st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
        st.markdown("<div style='margin-top: 1rem;'></div>", unsafe_allow_html=True)
        if st.button("✕ Fechar", key="close_stats", help="Fechar estatísticas"):
            st.session_state.show_stats = False
            st.rerun()
    
    st.markdown("---")
    if len(df_filtrado) == 0:
        st.warning("Nenhum resultado encontrado. Tente ajustar os filtros.")
    else:
        # Linha com resultados e botões
        col_results, col_stats, col_export = st.columns([2.5, 0.7, 0.7], gap="small")
        with col_results:
            st.markdown(f"#### Resultados: {len(df_filtrado)} registros")
        with col_stats:
            if st.button("📊 Estatísticas", key="btn_stats_main", use_container_width=True):
                st.session_state.show_stats = True
        with col_export:
            try:
                wb = criar_excel_glossario_formatado(df_filtrado)
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                st.download_button(
                    label="📥 Baixar dados em Excel",
                    data=buffer,
                    file_name=f"glossario_sinisa_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_glossario_{datetime.now().timestamp()}",
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"❌ Erro ao gerar Excel: {str(e)}")

        st.markdown("")
        if st.session_state.get("show_stats", False):
            exibir_estatisticas_modal(df_filtrado)
            st.markdown("---")
        
        # Tabela de resultados
        col1, col2, col3, col4, col5, col6, col7, col8 = st.columns([0.6, 3.0, 0.6, 0.6, 1.5, 0.6, 0.8, 0.4], gap="small", vertical_alignment="center")
        with col1:
            st.markdown('<div class="table-header">Código</div>', unsafe_allow_html=True)
        with col2:
            st.markdown('<div class="table-header">Nome da Informação</div>', unsafe_allow_html=True)
        with col3:
            st.markdown('<div class="table-header">Módulo</div>', unsafe_allow_html=True)
        with col4:
            st.markdown('<div class="table-header">Formulário</div>', unsafe_allow_html=True)
        with col5:
            st.markdown('<div class="table-header">Subformulário</div>', unsafe_allow_html=True)
        with col6:
            st.markdown('<div class="table-header">Unidade</div>', unsafe_allow_html=True)
        with col7:
            st.markdown('<div class="table-header">Área Resp.</div>', unsafe_allow_html=True)
        with col8:
            st.markdown('<div class="table-header">Ação</div>', unsafe_allow_html=True)
        
        for idx, row in df_filtrado.iterrows():
            col1, col2, col3, col4, col5, col6, col7, col8 = st.columns([0.6, 3.0, 0.6, 0.6, 1.5, 0.6, 0.8, 0.4], gap="small", vertical_alignment="center")
            with col1:
                st.markdown(f'<div class="table-row">{row["Código da informação"]}</div>', unsafe_allow_html=True)
            with col2:
                st.markdown(f'<div class="table-row">{row["Nome da informação"]}</div>', unsafe_allow_html=True)
            with col3:
                st.markdown(f'<div class="table-row">{row["Módulo"]}</div>', unsafe_allow_html=True)
            with col4:
                st.markdown(f'<div class="table-row">{row["Formulário"]}</div>', unsafe_allow_html=True)
            with col5:
                subform = row['Subformulário'] if pd.notna(row['Subformulário']) else '-'
                st.markdown(f'<div class="table-row">{subform}</div>', unsafe_allow_html=True)
            with col6:
                unidade = row['Unidade'] if pd.notna(row['Unidade']) else '-'
                st.markdown(f'<div class="table-row">{unidade}</div>', unsafe_allow_html=True)
            with col7:
                area = row['Área Responsável'] if pd.notna(row['Área Responsável']) else '-'
                st.markdown(f'<div class="table-row">{area}</div>', unsafe_allow_html=True)
            with col8:                    
                if st.button("🔍", key=f"detail_{idx}", help="Ver detalhes", use_container_width=True):
                    exibir_detalhes_popup(row)
            st.markdown('<div class="table-divider"></div>', unsafe_allow_html=True)

def pagina_glossario_indicadores():
    """Página Glossário - Indicadores"""
    st.markdown("""
    <div class="header-container">
        <h1 class="header-title">📚 SINISA</h1>
        <p class="header-subtitle">Sistema Nacional de Informações sobre Saneamento</p>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("### 📈 Indicadores")
    
    st.info("""
    ⏳ **Esta seção está em construção.**
    
    Em breve, você poderá acessar análises e indicadores detalhados do sistema SINISA.
    """)

def pagina_glossario_avisos():
    """Página Glossário - Avisos e Erros"""
    st.markdown("""
    <div class="header-container">
        <h1 class="header-title">📚 SINISA</h1>
        <p class="header-subtitle">Sistema Nacional de Informações sobre Saneamento</p>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("### ⚠️ Avisos e Erros")
    
    st.info("""
    ⏳ **Esta seção está em construção.**
    
    Em breve, você poderá acessar avisos e erros do sistema SINISA.
    """)

def pagina_relatorios_municipio():
    """Página Relatórios - Gerencial por Município"""
    st.markdown("""
    <div class="header-container">
        <h1 class="header-title">📚 SINISA</h1>
        <p class="header-subtitle">Sistema Nacional de Informações sobre Saneamento</p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("### 🏘️ Relatório Gerencial por Município")
    
    LARGURA_COD_DESC = 350
    LARGURA_UN = 60
    LARGURA_MES = 60
    
    # 1. Obter anos (abas) disponíveis usando a função existente
    anos_disponiveis = obter_anos_disponiveis()
        
    # Inicializar session state para o ano
    if "ano_filter_rel" not in st.session_state:
        st.session_state.ano_filter_rel = anos_disponiveis[0]
        
    # 2. Carregar dados do ano selecionado
    df_agems = load_dados_agems(st.session_state.ano_filter_rel)
    
    if df_agems is not None and len(df_agems) > 0:
        # Inicializar session state
        if "municipio_filter_rel" not in st.session_state:
            municipios_list = sorted(df_agems['municipio'].dropna().unique().tolist())
            # Colocar "ESTADO" na primeira posição
            if "ESTADO" in municipios_list:
                municipios_list.remove("ESTADO")
                municipios_list.insert(0, "ESTADO")
            st.session_state.municipio_filter_rel = municipios_list[0]
        if "modulo_filter_rel" not in st.session_state:
            st.session_state.modulo_filter_rel = "Todos"
        if "formulario_filter_rel" not in st.session_state:
            st.session_state.formulario_filter_rel = "Todos"
        if "subformulario_filter_rel" not in st.session_state:
            st.session_state.subformulario_filter_rel = "Todos"
        if "subgrupo_filter_rel" not in st.session_state:
            st.session_state.subgrupo_filter_rel = "Todos"
        if "search_unificado_rel" not in st.session_state:
            st.session_state.search_unificado_rel = ""
            
        # Callback para limpar filtros
        def limpar_filtros_callback():
            """Callback para limpar todos os filtros"""
            st.session_state.municipio_filter_rel = sorted(df_agems['municipio'].dropna().unique().tolist())[0]
            st.session_state.ano_filter_rel = anos_disponiveis[0]
            st.session_state.modulo_filter_rel = "Todos"
            st.session_state.formulario_filter_rel = "Todos"
            st.session_state.subformulario_filter_rel = "Todos"
            st.session_state.subgrupo_filter_rel = "Todos"
            st.session_state.search_unificado_rel = ""
            
        # Funções para obter valores únicos dos filtros
        @st.cache_data
        def get_municipios_rel():
            return sorted(df_agems['municipio'].dropna().unique().tolist())
            
        def get_filtro_relatorio(coluna, municipio, modulo="Todos", formulario="Todos", subformulario="Todos"):
            """Obtém valores únicos para filtros do relatório"""
            df_temp = df_agems[df_agems['municipio'] == municipio]
            if modulo != "Todos":
                df_temp = df_temp[df_temp['modulo'] == modulo]
            if formulario != "Todos":
                df_temp = df_temp[df_temp['formulario'] == formulario]
            if subformulario != "Todos":
                df_temp = df_temp[df_temp['subformulario_grupo'] == subformulario]
            return ["Todos"] + sorted(df_temp[coluna].dropna().unique().tolist())
            
        # Linha 1 - Município, Ano e botões
        col1, col2, col3, col4 = st.columns([1.5, 0.3, 0.5, 1.5])
        
        with col1:
            municipios_list = sorted(df_agems['municipio'].dropna().unique().tolist())
            if "ESTADO" in municipios_list:
                municipios_list.remove("ESTADO")
                municipios_list.insert(0, "ESTADO")

            municipio_selecionado = st.selectbox(
                "🏘️ Município",
                municipios_list,
                key="municipio_filter_rel"
            )
            
        with col2:
            ano_selecionado = st.selectbox(
                "📅 Ano",
                anos_disponiveis,
                key="ano_filter_rel"
            )
            
        with col3:
            st.markdown("<div style='padding-top: 10px;'></div>", unsafe_allow_html=True)
            st.button(
                "🔄 Limpar Filtros",
                use_container_width=True,
                key="btn_clear_rel",
                on_click=limpar_filtros_callback
            )
            
        with col4:
            st.markdown("<div style='padding-top: 10px;'></div>", unsafe_allow_html=True)
            export_placeholder = st.empty()
            
        # Linha 2 - Filtros
        col1, col2, col3, col4, col5 = st.columns([1, 1, 1, 1, 1.5])
        with col1:
            modulo_selecionado = st.selectbox(
                "Módulo",
                ["Todos"] + sort_modulos(get_filtro_relatorio('modulo', municipio_selecionado)[1:]),
                key="modulo_filter_rel"
            )
        with col2:
            formulario_selecionado = st.selectbox(
                "Formulário",
                get_filtro_relatorio('formulario', municipio_selecionado, modulo_selecionado),
                key="formulario_filter_rel"
            )
        with col3:
            subformulario_selecionado = st.selectbox(
                "Subformulário",
                get_filtro_relatorio('subformulario_grupo', municipio_selecionado, modulo_selecionado, formulario_selecionado),
                key="subformulario_filter_rel"
            )
        with col4:
            subgrupo_selecionado = st.selectbox(
                "Subgrupo",
                get_filtro_relatorio('subgrupo_palavra_chave', municipio_selecionado, modulo_selecionado, 
                                    formulario_selecionado, subformulario_selecionado),
                key="subgrupo_filter_rel"
            )
        with col5:
            busca_unificada = st.text_input(
                "🔎 Buscar (Código ou Descrição)",
                key="search_unificado_rel",
                placeholder="Digite código ou descrição"
            )
            
        # Filtrar dados
        df_municipio = df_agems[df_agems['municipio'] == municipio_selecionado].copy()
        df_municipio = aplicar_filtros(df_municipio, {
            'modulo': modulo_selecionado,
            'formulario': formulario_selecionado,
            'subformulario_grupo': subformulario_selecionado,
            'subgrupo_palavra_chave': subgrupo_selecionado
        })
        
        if busca_unificada:
            df_municipio = df_municipio[
                (df_municipio['codigo_da_informacao_indicador'].str.contains(busca_unificada, case=False, na=False)) |
                (df_municipio['nome_da_informacao_indicador'].str.contains(busca_unificada, case=False, na=False))
            ]
            
        # =====================================================================
        # INÍCIO DA SEÇÃO DE EXPORTAÇÃO ATUALIZADA COM POPUP
        # =====================================================================
        
        # Inicializar session state para ZIP
        if "zip_gerado" not in st.session_state:
            st.session_state.zip_gerado = None
        if "zip_pronto" not in st.session_state:
            st.session_state.zip_pronto = False
            
        # Função para o popup de geração do ZIP
        @st.dialog("📦 Baixar Todos os Municípios", width="large")
        def popup_gerar_zip():
            """Popup para gerar ZIP de todos os municípios"""
            try:
                # Placeholder para mensagem
                msg_placeholder = st.empty()
                
                # Mostrar mensagem de carregamento
                with msg_placeholder.container():
                    st.markdown("""
                    <div style="text-align: center; padding: 2rem;">
                        <h3 style="color: #1f4788; margin-bottom: 1rem;">⏳ Gerando ZIP com todos os municípios...</h3>
                        <p style="color: #666; font-size: 1rem;">Isso pode levar alguns segundos</p>
                    </div>
                    """, unsafe_allow_html=True)
                
                # Gerar ZIP
                zip_buffer = criar_zip_relatorios_todos_municipios(
                    df_agems,
                    ano_selecionado,
                    modulo_selecionado,
                    formulario_selecionado,
                    subformulario_selecionado,
                    subgrupo_selecionado,
                    busca_unificada
                )
                
                # Armazenar ZIP em session state
                st.session_state.zip_gerado = zip_buffer
                st.session_state.zip_pronto = True
                
                # Limpar placeholder e mostrar sucesso
                msg_placeholder.empty()
                
                st.markdown("""
                <div style="text-align: center; padding: 2rem;">
                    <h3 style="color: #2d7a4a; margin-bottom: 1.5rem;">✅ ZIP Gerado com Sucesso!</h3>
                </div>
                """, unsafe_allow_html=True)
                
                # Mostrar botão de download
                st.download_button(
                    label="📥 Clique aqui para Download",
                    data=st.session_state.zip_gerado,
                    file_name=f"relatorios_todos_municipios_{ano_selecionado}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                    mime="application/zip",
                    key=f"download_zip_popup_{datetime.now().timestamp()}",
                    use_container_width=True
                )
                
            except Exception as e:
                st.error(f"❌ Erro ao gerar ZIP: {str(e)}")
                
        # Botão de exportação
        with export_placeholder.container():
            col_btn1, col_btn2 = st.columns(2, gap="small")
            
            with col_btn1:
                try:
                    todas_colunas_meses = sorted([col for col in df_municipio.columns if '/' in col and len(col) == 7])
                    colunas_meses = [mes for mes in todas_colunas_meses if len(df_municipio[mes].dropna()) > 0]
                    
                    if len(df_municipio) > 0 and len(colunas_meses) > 0:
                        wb = criar_relatorio_excel_generico(df_municipio, municipio_selecionado, colunas_meses)
                        buffer = BytesIO()
                        wb.save(buffer)
                        buffer.seek(0)
                        st.download_button(
                            label="📥 Baixar Relatório em Excel",
                            data=buffer,
                            file_name=f"relatorio_{municipio_selecionado.replace(' ', '_')}_{ano_selecionado}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"download_relatorio_municipio_{datetime.now().timestamp()}",
                            use_container_width=True
                        )
                except Exception as e:
                    st.error(f"❌ Erro ao gerar relatório: {str(e)}")
            
            with col_btn2:
                # Botão para abrir popup
                if st.button("📦 Baixar Todos os Municípios", use_container_width=True, key="btn_gerar_zip_popup"):
                    popup_gerar_zip()
                    
        # =====================================================================
        # FIM DA SEÇÃO DE EXPORTAÇÃO
        # =====================================================================
                
        if len(df_municipio) > 0:
            # Obter colunas de meses
            todas_colunas_meses = sorted([col for col in df_municipio.columns if '/' in col and len(col) == 7])
            colunas_meses = [mes for mes in todas_colunas_meses if len(df_municipio[mes].dropna()) > 0]
            
            # Cabeçalho do Relatório
            st.markdown("---")
            st.markdown(f"""
            <div class="header-container">
                <h2 style="color: white; margin: 0;">📊 {municipio_selecionado} - {ano_selecionado}</h2>
                <p style="color: #e8f0ff; margin: 0.5rem 0 0 0;">Dados AGEMS - {datetime.now().strftime('%d/%m/%Y')}</p>
            </div>
            """, unsafe_allow_html=True)
            
            # Ordenar dados
            df_municipio = df_municipio.sort_values(['modulo', 'codigo_da_informacao_indicador'])
            modulos_unicos = sort_modulos(df_municipio['modulo'].unique().tolist())
            
            # Exibir dados por módulo
            for modulo in modulos_unicos:
                df_modulo = df_municipio[df_municipio['modulo'] == modulo]
                cores = get_modulo_color(modulo)
                
                st.markdown(f"""
                <div style="background-color: {cores['header_color']}; color: white; padding: 0.8rem; border-radius: 4px; margin-top: 1rem; margin-bottom: 0rem;">
                    <h3 style="margin: 0; font-size: 1.2rem;">{modulo}</h3>
                </div>
                """, unsafe_allow_html=True)
                
                # Se for "Informações Complementares", exibir sem agrupar por formulário
                if 'informações complementares' in modulo.lower():
                    html_tabela = criar_tabela_html(df_modulo, colunas_meses, LARGURA_COD_DESC, LARGURA_UN, LARGURA_MES)
                    st.markdown(html_tabela, unsafe_allow_html=True)
                    st.markdown("")
                else:
                    # Para Água e Esgoto, manter a estrutura original com formulário/subformulário
                    for formulario in df_modulo['formulario'].unique():
                        df_formulario = df_modulo[df_modulo['formulario'] == formulario]
                        
                        st.markdown(f"""
                        <div style="background-color: {cores['subheader_color']}; color: white; padding: 0.6rem 0.8rem; margin-top: 0rem; margin-bottom: 0rem; border-radius: 3px;">
                            <h4 style="margin: 0; font-size: 1rem;">{formulario}</h4>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        for subformulario in df_formulario['subformulario_grupo'].unique():
                            df_subformulario = df_formulario[df_formulario['subformulario_grupo'] == subformulario]
                            
                            st.markdown(f"""
                            <div style="background-color: {cores['subform_color']}; color: #1f4788; padding: 0.5rem 0.8rem; margin-top: 0rem; margin-bottom: 0rem; border-left: 3px solid {cores['border_color']}; font-weight: 600; font-size: 0.95rem;">
                                {subformulario}
                            </div>
                            """, unsafe_allow_html=True)
                            
                            for subgrupo in df_subformulario['subgrupo_palavra_chave'].unique():
                                df_subgrupo = df_subformulario[df_subformulario['subgrupo_palavra_chave'] == subgrupo]
                                
                                st.markdown(f"""
                                <div style="background-color: {cores['subgrupo_color']}; color: #1f4788; padding: 0.4rem 0.8rem; margin-top: 0rem; margin-bottom: 0rem; border-left: 4px solid {cores['border_color']}; font-weight: 600; font-size: 0.9rem;">
                                    {subgrupo}
                                </div>
                                """, unsafe_allow_html=True)
                                
                                html_tabela = criar_tabela_html(df_subgrupo, colunas_meses, LARGURA_COD_DESC, LARGURA_UN, LARGURA_MES)
                                st.markdown(html_tabela, unsafe_allow_html=True)
                                st.markdown("")
            st.markdown("---")
        else:
            st.warning(f"⚠️ Nenhum dado encontrado com os filtros selecionados")
    else:
        st.error("❌ Não foi possível carregar os dados AGEMS")

def pagina_relatorios_consolidado():
    """Página Relatórios - Consolidado"""
    st.markdown("""
    <div class="header-container">
        <h1 class="header-title">📚 SINISA</h1>
        <p class="header-subtitle">Sistema Nacional de Informações sobre Saneamento</p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("### 📊 Relatório Consolidado")
    
    # ===== OBTER ANOS DISPONÍVEIS DINAMICAMENTE =====
    anos_disponiveis = obter_anos_disponiveis()
    ano_padrao = anos_disponiveis[0]
    
    # Carregar dados do ano padrão
    df_sinisa = load_dados_sinisa(ano=ano_padrao)
    
    if df_sinisa is not None and len(df_sinisa) > 0:
        # ===== INICIALIZAR SESSION STATE =====
        if "modulo_filter_cons" not in st.session_state:
            st.session_state.modulo_filter_cons = "Todos"
        if "informacao_filter_cons" not in st.session_state:
            st.session_state.informacao_filter_cons = "Todos"
        if "formulario_filter_cons" not in st.session_state:
            st.session_state.formulario_filter_cons = "Todos"
        if "subformulario_filter_cons" not in st.session_state:
            st.session_state.subformulario_filter_cons = "Todos"
        if "subgrupo_filter_cons" not in st.session_state:
            st.session_state.subgrupo_filter_cons = "Todos"
        if "area_filter_cons" not in st.session_state:
            st.session_state.area_filter_cons = "Todos"
        if "indicadores_selecionados" not in st.session_state:
            st.session_state.indicadores_selecionados = []
        if "filtros_anteriores" not in st.session_state:
            st.session_state.filtros_anteriores = {
                "modulo": "Todos",
                "informacao": "Todos",
                "formulario": "Todos",
                "subformulario": "Todos",
                "subgrupo": "Todos",
                "area": "Todos"
            }
        if "ano_selecionado_cons" not in st.session_state:
            st.session_state.ano_selecionado_cons = ano_padrao
        if "visao_consolidado" not in st.session_state:
            st.session_state.visao_consolidado = "Acumulado no ano"
        if "incluir_colunas_vazias_cons" not in st.session_state:
            st.session_state.incluir_colunas_vazias_cons = False
            
        # Inicializar session state para os ZIPs gerados
        if "zip_modulo_gerado" not in st.session_state:
            st.session_state.zip_modulo_gerado = None
        if "zip_area_gerado" not in st.session_state:
            st.session_state.zip_area_gerado = None
        
        # Inicializar session state para Excel consolidado
        if "excel_consolidado_gerado" not in st.session_state:
            st.session_state.excel_consolidado_gerado = None
        
        # Callback para limpar filtros
        def limpar_filtros_consolidado_callback():
            """Callback para limpar todos os filtros do consolidado"""
            st.session_state.modulo_filter_cons = "Todos"
            st.session_state.informacao_filter_cons = "Todos"
            st.session_state.formulario_filter_cons = "Todos"
            st.session_state.subformulario_filter_cons = "Todos"
            st.session_state.subgrupo_filter_cons = "Todos"
            st.session_state.area_filter_cons = "Todos"
        
        # Funções para obter valores únicos dos filtros
        def get_filtro_consolidado(coluna, modulo="Todos", informacao="Todos", formulario="Todos", 
                                  subformulario="Todos", subgrupo="Todos"):
            """Obtém valores únicos para filtros do consolidado"""
            df_temp = df_sinisa.copy()
            if modulo != "Todos":
                df_temp = df_temp[df_temp['modulo'] == modulo]
            if informacao != "Todos":
                df_temp = df_temp[df_temp['informacao_indicador'] == informacao]
            if formulario != "Todos":
                df_temp = df_temp[df_temp['formulario'] == formulario]
            if subformulario != "Todos":
                df_temp = df_temp[df_temp['subformulario_grupo'] == subformulario]
            if subgrupo != "Todos":
                df_temp = df_temp[df_temp['subgrupo_palavra_chave'] == subgrupo]
            return ["Todos"] + sorted(df_temp[coluna].dropna().unique().tolist())
        
        # Linha 1 - Filtros
        col1, col2, col3, col4, col5, col6 = st.columns([1, 1, 1, 1, 1, 1])
        with col1:
            modulo_selecionado_cons = st.selectbox(
                "Módulo",
                ["Todos"] + sort_modulos(get_filtro_consolidado('modulo')[1:]),
                key="modulo_filter_cons"
            )
        with col2:
            informacao_selecionada_cons = st.selectbox(
                "Informação/Indicador",
                get_filtro_consolidado('informacao_indicador', modulo_selecionado_cons),
                key="informacao_filter_cons"
            )
        with col3:
            formulario_selecionado_cons = st.selectbox(
                "Formulário",
                get_filtro_consolidado('formulario', modulo_selecionado_cons, informacao_selecionada_cons),
                key="formulario_filter_cons"
            )
        with col4:
            subformulario_selecionado_cons = st.selectbox(
                "Subformulário",
                get_filtro_consolidado('subformulario_grupo', modulo_selecionado_cons, informacao_selecionada_cons, formulario_selecionado_cons),
                key="subformulario_filter_cons"
            )
        with col5:
            subgrupo_selecionado_cons = st.selectbox(
                "Subgrupo",
                get_filtro_consolidado('subgrupo_palavra_chave', modulo_selecionado_cons, informacao_selecionada_cons, 
                                      formulario_selecionado_cons, subformulario_selecionado_cons),
                key="subgrupo_filter_cons"
            )
        with col6:
            area_selecionada_cons = st.selectbox(
                "Área Responsável",
                get_filtro_consolidado('area_responsavel', modulo_selecionado_cons, informacao_selecionada_cons,
                                      formulario_selecionado_cons, subformulario_selecionado_cons, subgrupo_selecionado_cons),
                key="area_filter_cons"
            )
        
        # Filtrar dados
        df_consolidado = aplicar_filtros(df_sinisa, {
            'modulo': modulo_selecionado_cons,
            'informacao_indicador': informacao_selecionada_cons,
            'formulario': formulario_selecionado_cons,
            'subformulario_grupo': subformulario_selecionado_cons,
            'subgrupo_palavra_chave': subgrupo_selecionado_cons,
            'area_responsavel': area_selecionada_cons
        })
        
        # Remover linha do Estado
        df_consolidado = df_consolidado[df_consolidado['municipio'] != 'Estado']
        
        # Verificar se os filtros mudaram
        filtros_atuais = {
            "modulo": modulo_selecionado_cons,
            "informacao": informacao_selecionada_cons,
            "formulario": formulario_selecionado_cons,
            "subformulario": subformulario_selecionado_cons,
            "subgrupo": subgrupo_selecionado_cons,
            "area": area_selecionada_cons
        }
        if filtros_atuais != st.session_state.filtros_anteriores:
            st.session_state.filtros_anteriores = filtros_atuais.copy()
            st.session_state.indicadores_selecionados = sorted(df_consolidado['codigo_da_informacao_indicador'].unique().tolist())
        
        if len(df_consolidado) == 0:
            st.warning("⚠️ Nenhum dado encontrado com os filtros selecionados")
        else:               
            # Seção de seleção de indicadores
            st.markdown("""
            <div class="detail-section">
                <h3 style="color: #1f4788; margin-top: 0; margin-bottom: 1rem;">📊 Seleção de Informações/Indicadores</h3>
            </div>
            """, unsafe_allow_html=True)
            
            indicadores_disponiveis = sorted(df_consolidado['codigo_da_informacao_indicador'].unique().tolist())
            codigo_nome_unidade_map = {}
            # Obter TODOS os indicadores disponíveis (não apenas os que têm dados após filtros)
            indicadores_disponiveis = sorted(df_sinisa['codigo_da_informacao_indicador'].unique().tolist())

            # Criar mapa de código -> nome/unidade usando df_sinisa (completo)
            codigo_nome_unidade_map = {}
            for codigo in indicadores_disponiveis:
                # Procurar em df_sinisa (dados completos)
                row_data = df_sinisa[df_sinisa['codigo_da_informacao_indicador'] == codigo].iloc[0]
                codigo_nome_unidade_map[codigo] = formatar_nome_com_unidade(row_data['nome_da_informacao_indicador'], row_data['unidade'])
            
            indicadores_selecionados = st.multiselect(
                "Selecione as informações/indicadores para incluir no relatório",
                options=indicadores_disponiveis,
                default=st.session_state.indicadores_selecionados,
                format_func=lambda x: f"{x} - {codigo_nome_unidade_map.get(x, '')}",
                key="multiselect_indicadores"
            )
            st.session_state.indicadores_selecionados = indicadores_selecionados
            
            if len(indicadores_selecionados) > 0:
                df_consolidado_final = df_consolidado[df_consolidado['codigo_da_informacao_indicador'].isin(indicadores_selecionados)].copy()
            else:
                df_consolidado_final = df_consolidado.copy()           
            
            st.markdown("<div style='padding-top: 5px;'></div>", unsafe_allow_html=True)
            
            # ===== SELEÇÃO DE VISÃO LOGO ACIMA DA TABELA =====
            col_visao, col_checkbox = st.columns([4, 1])
            with col_visao:
                visao_consolidado = st.radio(
                    "Selecione a visão dos valores:",
                    options=[
                        "Acumulado no ano",
                        "Acumulado últimos 12 meses",
                        "Média - últimos 12 meses"
                    ],
                    horizontal=True,
                    key="visao_consolidado"
                )
            with col_checkbox:
                st.markdown("<div style='padding-top: 8px;'></div>", unsafe_allow_html=True)
                incluir_colunas_vazias = st.checkbox(
                    "Incluir colunas vazias",
                    value=False,
                    key="incluir_colunas_vazias_cons"
                )
            
            # Mapear visão para coluna
            mapa_visao_coluna = {
                "Acumulado no ano": "anual",
                "Acumulado últimos 12 meses": "acumulado_12_meses",
                "Média - últimos 12 meses": "media_12_meses"
            }
            coluna_valor_selecionada = mapa_visao_coluna[visao_consolidado]
            
            # ===== SELEÇÃO DE ANO (APARECE APENAS QUANDO "ACUMULADO NO ANO" É SELECIONADO) =====
            ano_para_filtro = ano_padrao  # Padrão
            if visao_consolidado == "Acumulado no ano":
                col_ano, col_espaco = st.columns([1, 5])
                with col_ano:
                    ano_para_filtro = st.selectbox(
                        "Selecione o ano:",
                        options=anos_disponiveis,
                        index=anos_disponiveis.index(st.session_state.ano_selecionado_cons) if st.session_state.ano_selecionado_cons in anos_disponiveis else 0,
                        key="ano_selecionado_cons"
                    )
                # Se o ano mudou, recarregar dados
                if ano_para_filtro != st.session_state.ano_selecionado_cons:
                    st.session_state.ano_selecionado_cons = ano_para_filtro
                    df_sinisa = load_dados_sinisa(ano=ano_para_filtro)
                    st.rerun()

            # =====================================================================
            # FUNÇÕES DE POPUP PARA EXPORTAÇÃO (SÓ EXECUTAM AO CLICAR)
            # =====================================================================
            
            # 1. Popup para Módulo
            @st.dialog("📦 Exportar por Formulário", width="large")
            def popup_exportar_modulo():
                try:
                    msg_placeholder = st.empty()
                    # Mensagem de carregamento
                    msg_placeholder.markdown("""
                    <div style="text-align: center; padding: 2rem;">
                        <h3 style="color: #1f4788; margin-bottom: 1rem;">⏳ Gerando ZIP por Módulo...</h3>
                        <p style="color: #666; font-size: 1rem;">Isso pode levar alguns segundos</p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    zip_buffer = criar_exportacao_zip_por_formulario(
                        tuple(df_sinisa.to_dict('records')),
                        coluna_valor=coluna_valor_selecionada,
                        incluir_colunas_vazias=incluir_colunas_vazias
                    )
                    
                    st.session_state.zip_modulo_gerado = zip_buffer
                    
                    # Substitui a mensagem de carregamento pela de sucesso
                    msg_placeholder.markdown("""
                    <div style="text-align: center; padding: 2rem;">
                        <h3 style="color: #2d7a4a; margin-bottom: 1.5rem;">✅ ZIP Gerado com Sucesso!</h3>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    st.download_button(
                        label="📥 Clique aqui para Download",
                        data=st.session_state.zip_modulo_gerado,
                        file_name=f"dados_por_formulario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                        mime="application/zip",
                        key=f"download_modulo_{datetime.now().timestamp()}",
                        use_container_width=True
                    )
                except Exception as e:
                    st.error(f"❌ Erro ao gerar exportação por módulo: {str(e)}")

            # 2. Popup para Área
            @st.dialog("📦 Exportar por Área", width="large")
            def popup_exportar_area():
                try:
                    msg_placeholder = st.empty()
                    # Mensagem de carregamento
                    msg_placeholder.markdown("""
                    <div style="text-align: center; padding: 2rem;">
                        <h3 style="color: #1f4788; margin-bottom: 1rem;">⏳ Gerando ZIP por Área...</h3>
                        <p style="color: #666; font-size: 1rem;">Isso pode levar alguns segundos</p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    zip_buffer = criar_exportacao_zip_por_area(
                        tuple(df_sinisa.to_dict('records')),
                        coluna_valor=coluna_valor_selecionada,
                        incluir_colunas_vazias=incluir_colunas_vazias
                    )
                    
                    st.session_state.zip_area_gerado = zip_buffer
                    
                    # Substitui a mensagem de carregamento pela de sucesso
                    msg_placeholder.markdown("""
                    <div style="text-align: center; padding: 2rem;">
                        <h3 style="color: #2d7a4a; margin-bottom: 1.5rem;">✅ ZIP Gerado com Sucesso!</h3>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    st.download_button(
                        label="📥 Clique aqui para Download",
                        data=st.session_state.zip_area_gerado,
                        file_name=f"dados_por_area_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                        mime="application/zip",
                        key=f"download_area_{datetime.now().timestamp()}",
                        use_container_width=True
                    )
                except Exception as e:
                    st.error(f"❌ Erro ao gerar exportação por área: {str(e)}")

            # 3. Popup para Excel Consolidado
            @st.dialog("📊 Exportar em Excel", width="large")
            def popup_exportar_excel():
                try:
                    msg_placeholder = st.empty()
                    # Mensagem de carregamento
                    msg_placeholder.markdown("""
                    <div style="text-align: center; padding: 2rem;">
                        <h3 style="color: #1f4788; margin-bottom: 1rem;">⏳ Gerando arquivo Excel...</h3>
                        <p style="color: #666; font-size: 1rem;">Isso pode levar alguns segundos</p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    wb = criar_relatorio_consolidado_excel(
                        df_consolidado_final, 
                        tuple(indicadores_selecionados), 
                        coluna_valor=coluna_valor_selecionada,
                        incluir_colunas_vazias=incluir_colunas_vazias
                    )
                    buffer = BytesIO()
                    wb.save(buffer)
                    buffer.seek(0)
                    
                    st.session_state.excel_consolidado_gerado = buffer
                    
                    # Substitui a mensagem de carregamento pela de sucesso
                    msg_placeholder.markdown("""
                    <div style="text-align: center; padding: 2rem;">
                        <h3 style="color: #2d7a4a; margin-bottom: 1.5rem;">✅ Excel Gerado com Sucesso!</h3>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    st.download_button(
                        label="📥 Clique aqui para Download",
                        data=st.session_state.excel_consolidado_gerado,
                        file_name=f"relatorio_consolidado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_relatorio_consolidado_{datetime.now().timestamp()}",
                        use_container_width=True
                    )
                except Exception as e:
                    st.error(f"❌ Erro ao gerar relatório: {str(e)}")

            # Linha de botões
            col1, col2, col3, col4 = st.columns([1, 1, 1, 1], gap="small")
            
            with col1:
                st.button(
                    "🔄 Limpar Filtros",
                    key="btn_clear_cons",
                    on_click=limpar_filtros_consolidado_callback,
                    use_container_width=True
                )
            
            # BOTÃO EXCEL - ABRE POPUP AO CLICAR
            with col2:
                if st.button("📥 Exportar em Excel", use_container_width=True, key="btn_excel_cons"):
                    popup_exportar_excel()
            
            with col3:
                if st.button("📦 Exportar por Formulário", use_container_width=True, key="btn_modulo_cons"):
                    popup_exportar_modulo()
            
            with col4:
                if st.button("📦 Exportar por Área", use_container_width=True, key="btn_area_cons"):
                    popup_exportar_area()
            
            st.markdown("---")          

            # Exibir tabela consolidada
            html_tabela_consolidada = criar_tabela_consolidada_html(
                df_consolidado_final, 
                indicadores_selecionados,
                coluna_valor_selecionada,
                incluir_colunas_vazias
            )
            st.markdown(html_tabela_consolidada, unsafe_allow_html=True)
            st.markdown("---")
    else:
        st.error(f"❌ Não foi possível carregar os dados SINISA do ano {ano_padrao}")

def pagina_relatorios_verificacao():
    """Página Relatórios - Verificação"""
    st.markdown("""
    <div class="header-container">
        <h1 class="header-title">📚 SINISA</h1>
        <p class="header-subtitle">Sistema Nacional de Informações sobre Saneamento</p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("### ✅ Verificação")
    
    # Carregar dados
    df_glossario = load_data()
    df_correspondencia = load_correspondencia_sigis_sinisa()
    df_dados_sigis = load_dados_sigis()
    df_sinisa = load_dados_sinisa()  # ← FUNÇÃO CORRETA
    
    if df_glossario is None or df_correspondencia is None or df_dados_sigis is None or df_sinisa is None:
        st.error("❌ Não foi possível carregar os dados necessários")
        return
    
    # TRATAMENTO ROBUSTO DE DADOS
    df_correspondencia['codigo_sinisa'] = df_correspondencia['codigo_sinisa'].astype(str).str.strip()
    df_correspondencia['termos_sinisa'] = df_correspondencia['termos_sinisa'].astype(str).str.strip()
    
    def limpar_codigo_sigis(val):
        if pd.isna(val) or val == "" or str(val).strip() == "nan":
            return ""
        val_str = str(val).strip()
        if val_str.endswith('.0'):
            return val_str[:-2]
        return val_str
        
    df_correspondencia['termos_sigis'] = df_correspondencia['termos_sigis'].apply(limpar_codigo_sigis)
    df_dados_sigis['cod_sigis'] = df_dados_sigis['cod_sigis'].apply(limpar_codigo_sigis)
    
    # Funções de filtro
    @st.cache_data
    def get_informacoes_verificacao():
        """Obtém lista de informações únicas do glossário com código + descrição"""
        informacoes_dict = {}
        for _, row in df_glossario.iterrows():
            if pd.notna(row['Código da informação']):
                codigo = str(row['Código da informação']).strip()
                nome = str(row['Nome da informação']).strip()
                informacoes_dict[f"{codigo} - {nome}"] = codigo
        
        opcoes = ["Selecione uma informação"] + sorted(informacoes_dict.keys())
        return opcoes, informacoes_dict
    
    @st.cache_data
    def get_municipios_verificacao():
        """Obtém lista de municípios"""
        municipios = sorted(df_dados_sigis['nome_localidade'].dropna().unique().tolist())
        return municipios
    
    def get_descritivo_informacao(codigo_informacao):
        """Obtém descritivo completo da informação"""
        df_info = df_glossario[df_glossario['Código da informação'].astype(str).str.strip() == codigo_informacao]
        if len(df_info) == 0:
            return None
        return df_info.iloc[0]
    
    def get_componentes_informacao(codigo_selecionado):
        """Obtém os componentes (termos) de uma informação SINISA"""
        # 1. Tentar buscar como código principal (ex: GFI1003)
        df_comp_principal = df_correspondencia[df_correspondencia['codigo_sinisa'] == codigo_selecionado]
        
        if len(df_comp_principal) > 0:
            return df_comp_principal[['termos_sinisa', 'termos_sigis']].drop_duplicates().values.tolist()
            
        # 2. Se não encontrou, tentar buscar como termo (ex: GTA1209)
        df_comp_termo = df_correspondencia[df_correspondencia['termos_sinisa'] == codigo_selecionado]
        
        if len(df_comp_termo) > 0:
            return df_comp_termo[['termos_sinisa', 'termos_sigis']].drop_duplicates().values.tolist()
            
        return []
    
    def get_descricao_termo(termo_sinisa):
        """Obtém descrição de um termo SINISA"""
        df_termo = df_glossario[df_glossario['Código da informação'].astype(str).str.strip() == termo_sinisa]
        if len(df_termo) == 0:
            return ""
        return str(df_termo.iloc[0]['Nome da informação']).strip()
    
    def criar_tabela_verificacao(codigo_sinisa, municipio_selecionado, ano_selecionado):
        """Cria tabela com componentes e valores dos 12 meses do ano selecionado"""
        # Obter componentes
        componentes = get_componentes_informacao(codigo_sinisa)
        
        if not componentes:
            st.warning(f"Nenhum componente encontrado para {codigo_sinisa} na tabela de correspondência.")
            return None, None
        
        # Filtrar dados SIGIS para o município e ano
        df_municipio = df_dados_sigis[
            (df_dados_sigis['nome_localidade'] == municipio_selecionado)
        ].copy()
        
        if len(df_municipio) == 0:
            st.warning(f"Nenhum dado encontrado para o município {municipio_selecionado}.")
            return None, None
        
        # Converter data para datetime
        df_municipio['data'] = pd.to_datetime(df_municipio['data'])
        
        # Filtrar pelo ano selecionado
        df_municipio = df_municipio[df_municipio['data'].dt.year == ano_selecionado]
        
        # Criar lista de meses (01/2025, 02/2025, ..., 12/2025)
        meses_formatados = [f"{i:02d}/{ano_selecionado}" for i in range(1, 13)]
        
        # Informação principal
        info_principal = get_descritivo_informacao(codigo_sinisa)
        descricao_principal = info_principal['Nome da informação'] if info_principal is not None else ""
        
        # Construir tabela HTML
        html = '<div style="overflow-x: auto; margin: 1rem 0; border-radius: 4px; border: 1px solid #d0d8e0;">'
        html += '<table style="width: 100%; border-collapse: collapse; font-size: 0.9rem;">'
        
        # Cabeçalho da tabela
        html += '<thead><tr style="background-color: #0B3040; color: white;">'
        html += '<th style="padding: 0.5rem 0.75rem; text-align: left; border: 1px solid #d0d8e0; font-weight: 600; min-width: 500px; height: 40px;">Informação</th>'
        for mes in meses_formatados:
            html += f'<th style="padding: 0.5rem 0.75rem; text-align: center; border: 1px solid #d0d8e0; font-weight: 600; white-space: nowrap; height: 40px;">{mes}</th>'
        html += '<th style="padding: 0.5rem 0.75rem; text-align: center; border: 1px solid #d0d8e0; font-weight: 600; white-space: nowrap; height: 40px;">Acumulado</th>'
        html += '</tr></thead>'
        
        # Dados
        html += '<tbody>'
        
        # Armazenar dados para o gráfico e download
        dados_grafico = {
            'informacao': [],
            'descricao': [],
            'mes': [],
            'mes_formatado': [],
            'valor': [],
            'nivel': [],
            'cor': [],
            'acumulado': []
        }
        
        # Dicionário de cores por nível
        cores_por_nivel = {
            'principal': '#1f4788',
            'nivel1': '#5a8ac4',
            'sigis': '#a8c5e8'
        }
        
        # Linha 1: Informação Principal (GFI1003 ou GTA1209) - AZUL CLARO
        html += f'<tr style="background-color: #ADD8E6;">'
        html += f'<td style="padding: 0.75rem; border: 1px solid #d0d8e0; font-weight: 600; word-wrap: break-word; white-space: normal;">{codigo_sinisa} - {descricao_principal}</td>'
        
        # Valores para a informação principal (soma dos componentes)
        valor_acumulado_sinisa = 0
        for mes_num in range(1, 13):
            valor_total = 0
            for termo_sinisa, termo_sigis in componentes:
                if termo_sigis:
                    df_termo = df_municipio[df_municipio['cod_sigis'] == termo_sigis]
                    df_mes = df_termo[df_termo['data'].dt.month == mes_num]
                    if len(df_mes) > 0:
                        try:
                            valor_total += float(df_mes['valor'].values[0])
                        except:
                            pass
            
            valor_formatado = f"{valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if valor_total > 0 else "-"
            html += f'<td style="padding: 0.75rem; border: 1px solid #d0d8e0; text-align: right; font-weight: 600;">{valor_formatado}</td>'
            
            # Adicionar ao gráfico
            dados_grafico['informacao'].append(codigo_sinisa)
            dados_grafico['descricao'].append(f"{codigo_sinisa} - {descricao_principal}")
            dados_grafico['mes'].append(f"{mes_num:02d}/{ano_selecionado}")
            dados_grafico['mes_formatado'].append(f"{mes_num:02d}/{ano_selecionado}")
            dados_grafico['valor'].append(valor_total)
            dados_grafico['nivel'].append('principal')
            dados_grafico['cor'].append(cores_por_nivel['principal'])
            dados_grafico['acumulado'].append(0)  # Será preenchido depois
        
        # Coluna Acumulado - BUSCAR DO SINISA
        # Filtrar SINISA pelo código, município e ano
        df_sinisa_filtrado = df_sinisa[
            (df_sinisa['codigo_da_informacao_indicador'] == codigo_sinisa) &
            (df_sinisa['municipio'] == municipio_selecionado)
        ]
        
        # Obter o valor da coluna "anual" do SINISA
        if len(df_sinisa_filtrado) > 0:
            valor_acumulado_sinisa = df_sinisa_filtrado.iloc[0]['anual']
            if pd.notna(valor_acumulado_sinisa):
                try:
                    valor_acumulado_sinisa = float(valor_acumulado_sinisa)
                    valor_acum_formatado = f"{valor_acumulado_sinisa:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                except:
                    valor_acum_formatado = "-"
                    valor_acumulado_sinisa = 0
            else:
                valor_acum_formatado = "-"
                valor_acumulado_sinisa = 0
        else:
            valor_acum_formatado = "-"
            valor_acumulado_sinisa = 0
        
        html += f'<td style="padding: 0.75rem; border: 1px solid #d0d8e0; text-align: right; font-weight: 600;">{valor_acum_formatado}</td>'
        html += '</tr>'
        
        # Atualizar acumulado na linha principal
        if len(dados_grafico['acumulado']) >= 12:
            for i in range(12):
                dados_grafico['acumulado'][i] = valor_acumulado_sinisa
        
        # Verificar se a informação selecionada é ela mesma um termo (como GTA1209)
        is_termo_direto = all(t_sinisa == codigo_sinisa for t_sinisa, _ in componentes)
        
        # Se for um termo direto (GTA1209), não precisamos repetir a linha do SINISA, vamos direto para os SIGIS
        if is_termo_direto:
            for idx, (termo_sinisa, termo_sigis) in enumerate(componentes):
                
                if termo_sigis:
                    df_sigis_desc = df_municipio[df_municipio['cod_sigis'] == termo_sigis]
                    desc_sigis = df_sigis_desc['desc_sigis'].values[0] if len(df_sigis_desc) > 0 else f"{termo_sigis}"
                    
                    # BRANCO para SIGIS
                    html += f'<tr style="background-color: #FFFFFF;">'
                    html += f'<td style="padding: 0.75rem; border: 1px solid #d0d8e0; font-weight: 400; padding-left: 2rem; color: #333; font-size: 0.85rem; word-wrap: break-word; white-space: normal;">SIGIS: {desc_sigis}</td>'
                    
                    valor_acum_sigis = 0
                    for mes_num in range(1, 13):
                        df_mes = df_sigis_desc[df_sigis_desc['data'].dt.month == mes_num]
                        valor = df_mes['valor'].values[0] if len(df_mes) > 0 else "-"
                        
                        if valor != "-":
                            try:
                                valor_float = float(valor)
                                valor_formatado = f"{valor_float:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                                # Capturar acumulado de dezembro do SIGIS
                                if mes_num == 12:
                                    valor_acum_sigis = float(df_mes['valor_acumulado'].values[0]) if 'valor_acumulado' in df_mes.columns else valor_float
                            except:
                                valor_formatado = str(valor)
                        else:
                            valor_formatado = "-"
                        
                        html += f'<td style="padding: 0.75rem; border: 1px solid #d0d8e0; text-align: right;">{valor_formatado}</td>'
                        
                        # Adicionar ao gráfico
                        if valor != "-":
                            dados_grafico['informacao'].append(f"SIGIS: {termo_sigis}")
                            dados_grafico['descricao'].append(f"SIGIS: {desc_sigis}")
                            dados_grafico['mes'].append(f"{mes_num:02d}/{ano_selecionado}")
                            dados_grafico['mes_formatado'].append(f"{mes_num:02d}/{ano_selecionado}")
                            dados_grafico['valor'].append(float(valor))
                            dados_grafico['nivel'].append('sigis')
                            dados_grafico['cor'].append(cores_por_nivel['sigis'])
                            dados_grafico['acumulado'].append(valor_acum_sigis if mes_num == 12 else 0)
                    
                    # Coluna Acumulado - USAR valor_acumulado DO SIGIS
                    valor_acum_fmt = f"{valor_acum_sigis:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if valor_acum_sigis > 0 else "-"
                    html += f'<td style="padding: 0.75rem; border: 1px solid #d0d8e0; text-align: right;">{valor_acum_fmt}</td>'
                    html += '</tr>'
        
        # Se for uma informação pai (GFI1003), mostramos a hierarquia completa (SINISA -> SIGIS)
        else:
            # Agrupar por termo_sinisa para não repetir a linha do SINISA se houver múltiplos SIGIS
            termos_agrupados = {}
            for t_sinisa, t_sigis in componentes:
                if t_sinisa not in termos_agrupados:
                    termos_agrupados[t_sinisa] = []
                if t_sigis:
                    termos_agrupados[t_sinisa].append(t_sigis)
            
            for termo_sinisa, lista_sigis in termos_agrupados.items():
                descricao_termo = get_descricao_termo(termo_sinisa)
                
                # CINZA para componentes de nível 1 (SINISA)
                html += f'<tr style="background-color: #D3D3D3;">'
                html += f'<td style="padding: 0.75rem; border: 1px solid #d0d8e0; font-weight: 500; padding-left: 2rem; word-wrap: break-word; white-space: normal;">{termo_sinisa} - {descricao_termo}</td>'
                
                # Valores por mês (soma dos SIGIS filhos)
                valor_acum_termo = 0
                for mes_num in range(1, 13):
                    valor_total_termo = 0
                    for t_sigis in lista_sigis:
                        df_termo = df_municipio[df_municipio['cod_sigis'] == t_sigis]
                        df_mes = df_termo[df_termo['data'].dt.month == mes_num]
                        if len(df_mes) > 0:
                            try:
                                valor_total_termo += float(df_mes['valor'].values[0])
                                # Capturar acumulado de dezembro
                                if mes_num == 12:
                                    valor_acum_termo += float(df_mes['valor_acumulado'].values[0])
                            except:
                                pass
                    
                    valor_formatado = f"{valor_total_termo:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if valor_total_termo > 0 else "-"
                    html += f'<td style="padding: 0.75rem; border: 1px solid #d0d8e0; text-align: right;">{valor_formatado}</td>'
                    
                    # Adicionar ao gráfico
                    dados_grafico['informacao'].append(termo_sinisa)
                    dados_grafico['descricao'].append(f"{termo_sinisa} - {descricao_termo}")
                    dados_grafico['mes'].append(f"{mes_num:02d}/{ano_selecionado}")
                    dados_grafico['mes_formatado'].append(f"{mes_num:02d}/{ano_selecionado}")
                    dados_grafico['valor'].append(valor_total_termo)
                    dados_grafico['nivel'].append('nivel1')
                    dados_grafico['cor'].append(cores_por_nivel['nivel1'])
                    dados_grafico['acumulado'].append(valor_acum_termo if mes_num == 12 else 0)
                
                # Coluna Acumulado - USAR SIGIS
                valor_acum_fmt = f"{valor_acum_termo:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if valor_acum_termo > 0 else "-"
                html += f'<td style="padding: 0.75rem; border: 1px solid #d0d8e0; text-align: right;">{valor_acum_fmt}</td>'
                html += '</tr>'
                
                # Linhas dos filhos SIGIS (ex: 8747) - BRANCO
                for t_sigis in lista_sigis:
                    df_sigis_desc = df_municipio[df_municipio['cod_sigis'] == t_sigis]
                    desc_sigis = df_sigis_desc['desc_sigis'].values[0] if len(df_sigis_desc) > 0 else f"{t_sigis}"
                    
                    html += f'<tr style="background-color: #FFFFFF;">'
                    html += f'<td style="padding: 0.75rem; border: 1px solid #d0d8e0; font-weight: 400; padding-left: 4rem; color: #333; font-size: 0.85rem; word-wrap: break-word; white-space: normal;">SIGIS: {desc_sigis}</td>'
                    
                    valor_acum_sigis = 0
                    for mes_num in range(1, 13):
                        df_mes = df_sigis_desc[df_sigis_desc['data'].dt.month == mes_num]
                        valor = df_mes['valor'].values[0] if len(df_mes) > 0 else "-"
                        
                        if valor != "-":
                            try:
                                valor_float = float(valor)
                                valor_formatado = f"{valor_float:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                                # Capturar acumulado de dezembro
                                if mes_num == 12:
                                    valor_acum_sigis = float(df_mes['valor_acumulado'].values[0]) if 'valor_acumulado' in df_mes.columns else valor_float
                            except:
                                valor_formatado = str(valor)
                        else:
                            valor_formatado = "-"
                        
                        html += f'<td style="padding: 0.75rem; border: 1px solid #d0d8e0; text-align: right;">{valor_formatado}</td>'
                        
                        # Adicionar ao gráfico
                        if valor != "-":
                            dados_grafico['informacao'].append(f"SIGIS: {t_sigis}")
                            dados_grafico['descricao'].append(f"SIGIS: {desc_sigis}")
                            dados_grafico['mes'].append(f"{mes_num:02d}/{ano_selecionado}")
                            dados_grafico['mes_formatado'].append(f"{mes_num:02d}/{ano_selecionado}")
                            dados_grafico['valor'].append(float(valor))
                            dados_grafico['nivel'].append('sigis')
                            dados_grafico['cor'].append(cores_por_nivel['sigis'])
                            dados_grafico['acumulado'].append(valor_acum_sigis if mes_num == 12 else 0)
                    
                    # Coluna Acumulado - USAR SIGIS
                    valor_acum_fmt = f"{valor_acum_sigis:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if valor_acum_sigis > 0 else "-"
                    html += f'<td style="padding: 0.75rem; border: 1px solid #d0d8e0; text-align: right;">{valor_acum_fmt}</td>'
                    html += '</tr>'
        
        html += '</tbody></table></div>'
        st.markdown(html, unsafe_allow_html=True)
        
        return dados_grafico, html
    
    def gerar_excel_formatado(dados_grafico, html_tabela, codigo_sinisa, municipio_selecionado, ano_selecionado):
        """Gera Excel com a mesma formatação e estilo da tabela"""
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Verificação"
        
        # Obter dados únicos
        df_dados = pd.DataFrame(dados_grafico)
        
        # Cabeçalho - USAR mes_formatado (MM/AAAA)
        meses_unicos = sorted(df_dados['mes_formatado'].unique())
        
        # Linha 1: Cabeçalho
        ws['A1'] = "Informação"
        for idx, mes in enumerate(meses_unicos, start=2):
            ws.cell(row=1, column=idx).value = mes
        ws.cell(row=1, column=len(meses_unicos) + 2).value = "Anual"
        
        # Estilos
        header_fill = PatternFill(start_color="0B3040", end_color="0B3040", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        azul_claro_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        cinza_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        branco_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        border = Border(
            left=Side(style='thin', color='d0d8e0'),
            right=Side(style='thin', color='d0d8e0'),
            top=Side(style='thin', color='d0d8e0'),
            bottom=Side(style='thin', color='d0d8e0')
        )
        
        # Aplicar estilo ao cabeçalho
        for col in range(1, len(meses_unicos) + 3):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Dados
        row_num = 2
        informacoes_unicas = df_dados['descricao'].unique()
        
        for desc in informacoes_unicas:
            df_info = df_dados[df_dados['descricao'] == desc]
            nivel = df_info['nivel'].values[0]
            
            # Determinar cor baseado no nível
            if nivel == 'principal':
                fill = azul_claro_fill
                font = Font(bold=True, size=11)
            elif nivel == 'nivel1':
                fill = cinza_fill
                font = Font(bold=True, size=10)
            else:  # sigis
                fill = branco_fill
                font = Font(size=10)
            
            # Coluna de informação - DESCRIÇÃO COMPLETA
            ws.cell(row=row_num, column=1).value = desc
            ws.cell(row=row_num, column=1).fill = fill
            ws.cell(row=row_num, column=1).font = font
            ws.cell(row=row_num, column=1).border = border
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Valores por mês - USAR mes_formatado
            for idx, mes in enumerate(meses_unicos, start=2):
                valor = df_info[df_info['mes_formatado'] == mes]['valor'].values
                if len(valor) > 0:
                    cell_value = valor[0]
                    ws.cell(row=row_num, column=idx).value = cell_value
                    ws.cell(row=row_num, column=idx).number_format = '#,##0.00'
                else:
                    ws.cell(row=row_num, column=idx).value = 0
                
                cell = ws.cell(row=row_num, column=idx)
                cell.fill = fill
                cell.font = font
                cell.border = border
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Coluna Acumulado - USAR valor_acumulado
            acumulado_valor = df_info['acumulado'].values[0] if len(df_info['acumulado'].values) > 0 else 0
            ws.cell(row=row_num, column=len(meses_unicos) + 2).value = acumulado_valor
            cell = ws.cell(row=row_num, column=len(meses_unicos) + 2)
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.number_format = '#,##0.00'
            
            row_num += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 80
        for idx in range(2, len(meses_unicos) + 3):
            ws.column_dimensions[get_column_letter(idx)].width = 15
        
        # Salvar em buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    # Interface
    col1, col2 = st.columns([2, 2])
    
    # Obter opções de informações
    opcoes_info, info_dict = get_informacoes_verificacao()
    
    with col1:
        informacao_selecionada_display = st.selectbox(
            "Selecione a Informação/Indicador",
            opcoes_info,
            key="informacao_verif_select",
            index=0
        )
        # Extrair código da seleção
        if informacao_selecionada_display != "Selecione uma informação":
            informacao_selecionada = info_dict[informacao_selecionada_display]
        else:
            informacao_selecionada = None
    
    with col2:
        municipio_selecionado = st.selectbox(
            "Selecione o Município",
            get_municipios_verificacao(),
            key="municipio_verif_select"
        )
    
    st.markdown("---")
    
    # Exibir descritivo da informação selecionada
    if informacao_selecionada is not None:
        row = get_descritivo_informacao(informacao_selecionada)
        
        if row is not None:
            st.markdown("#### 📖 Detalhes da Informação")
            exibir_detalhes_modal(row, "verif_0")
            st.markdown("---")
            
            # Tabela de verificação - LAYOUT COM ANO E BOTÃO
            col_titulo, col_ano, col_botao = st.columns([1.5, 0.8, 0.7], vertical_alignment="center")
            
            with col_titulo:
                st.markdown("#### 📊 Composição da Informação")
            
            with col_ano:
                # Obter anos disponíveis
                anos_disponiveis = sorted(df_dados_sigis['data'].dt.year.unique(), reverse=True)
                ano_selecionado = st.selectbox(
                    "Ano",
                    anos_disponiveis,
                    key="ano_verif_select",
                    label_visibility="collapsed"
                )
            
            with col_botao:
                pass
            
            dados_grafico, html_tabela = criar_tabela_verificacao(informacao_selecionada, municipio_selecionado, ano_selecionado)
            
            if dados_grafico is not None:
                # Botão de download alinhado à direita
                with col_botao:
                    # Gerar Excel formatado
                    buffer_excel = gerar_excel_formatado(dados_grafico, html_tabela, informacao_selecionada, municipio_selecionado, ano_selecionado)
                    
                    st.download_button(
                        label="📥 Baixar",
                        data=buffer_excel,
                        file_name=f"verificacao_{informacao_selecionada}_{municipio_selecionado}_{ano_selecionado}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                
                st.markdown("---")
                
                # ============================================================
                # GRÁFICO: COLUNAS SIGIS + LINHAS HIERARQUIA
                # ============================================================
                st.markdown("#### 📈 Gráfico de Evolução")
                
                df_grafico = pd.DataFrame(dados_grafico)
                
                # Separar dados
                df_sigis = df_grafico[df_grafico['nivel'] == 'sigis'].copy()
                df_nivel1 = df_grafico[df_grafico['nivel'] == 'nivel1'].copy()
                df_principal = df_grafico[df_grafico['nivel'] == 'principal'].copy()
                
                # Obter valores mín e máx para escala uniforme
                todos_valores = pd.concat([df_sigis['valor'], df_nivel1['valor'], df_principal['valor']])
                valor_min = todos_valores.min()
                valor_max = todos_valores.max()
                
                # Criar figura com eixos secundários
                fig1 = make_subplots(specs=[[{"secondary_y": True}]])
                
                # Paleta de cores para SIGIS com transparência
                paleta_cores = [
                    'rgba(255, 107, 107, 0.6)',      # Vermelho coral
                    'rgba(78, 205, 196, 0.6)',       # Turquesa
                    'rgba(69, 183, 209, 0.6)',       # Azul céu
                    'rgba(255, 160, 122, 0.6)',      # Salmão claro
                    'rgba(152, 216, 200, 0.6)',      # Menta
                    'rgba(247, 220, 111, 0.6)',      # Amarelo suave
                    'rgba(187, 143, 206, 0.6)',      # Roxo suave
                    'rgba(133, 193, 226, 0.6)',      # Azul claro
                    'rgba(248, 184, 139, 0.6)',      # Pêssego
                    'rgba(169, 223, 191, 0.6)',      # Verde suave
                    'rgba(241, 148, 138, 0.6)',      # Rosa coral
                    'rgba(215, 189, 226, 0.6)'       # Lavanda
                ]
                
                # Adicionar colunas empilhadas SIGIS
                sigis_unicos = df_sigis['descricao'].unique()
                for idx, sigis in enumerate(sigis_unicos):
                    df_sigis_item = df_sigis[df_sigis['descricao'] == sigis]
                    cor = paleta_cores[idx % len(paleta_cores)]
                    
                    fig1.add_trace(
                        go.Bar(
                            x=df_sigis_item['mes_formatado'],
                            y=df_sigis_item['valor'],
                            name=sigis,
                            marker=dict(color=cor),
                            hovertemplate='<b>%{fullData.name}</b><br>Valor: %{y:,.2f}<extra></extra>',
                            showlegend=True
                        ),
                        secondary_y=False
                    )
                
                # Adicionar linhas para Nível 1 (SINISA intermediários)
                nivel1_unicos = df_nivel1['descricao'].unique()
                for nivel1 in nivel1_unicos:
                    df_nivel1_item = df_nivel1[df_nivel1['descricao'] == nivel1]
                    
                    fig1.add_trace(
                        go.Scatter(
                            x=df_nivel1_item['mes_formatado'],
                            y=df_nivel1_item['valor'],
                            mode='lines',
                            name=nivel1,
                            line=dict(color='rgba(90, 138, 196, 0.8)', width=2, dash='dash', shape='spline'),
                            hovertemplate='<b>%{fullData.name}</b><br>Valor: %{y:,.2f}<extra></extra>',
                            showlegend=True
                        ),
                        secondary_y=True
                    )
                
                # Adicionar linha para Principal
                fig1.add_trace(
                    go.Scatter(
                        x=df_principal['mes_formatado'],
                        y=df_principal['valor'],
                        mode='lines',
                        name=f"{informacao_selecionada} (Total)",
                        line=dict(color='rgba(31, 71, 136, 1)', width=3, shape='spline'),
                        hovertemplate='<b>%{fullData.name}</b><br>Valor: %{y:,.2f}<extra></extra>',
                        showlegend=True
                    ),
                    secondary_y=True
                )
                
                fig1.update_layout(
                    title=f"Composição de {informacao_selecionada} - {municipio_selecionado} ({ano_selecionado})",
                    xaxis_title="",
                    barmode='stack',
                    hovermode='x unified',
                    height=650,
                    template='plotly_white',
                    font=dict(size=11, family="Arial"),
                    showlegend=True,
                    legend=dict(
                        orientation="v",
                        yanchor="top",
                        y=0.99,
                        xanchor="right",
                        x=0.99,
                        bgcolor="rgba(255, 255, 255, 0.9)",
                        bordercolor="rgba(0, 0, 0, 0.2)",
                        borderwidth=1,
                        font=dict(size=10)
                    ),
                    margin=dict(l=60, r=60, t=100, b=60),
                    plot_bgcolor='white',
                    paper_bgcolor='white'
                )
                
                # Configurar eixos Y com mesma escala e SEM VALORES
                fig1.update_yaxes(
                    range=[valor_min * 0.95, valor_max * 1.05],
                    secondary_y=False,
                    gridcolor='rgba(200, 200, 200, 0.2)',
                    showticklabels=False
                )
                fig1.update_yaxes(
                    range=[valor_min * 0.95, valor_max * 1.05],
                    secondary_y=True,
                    gridcolor='rgba(200, 200, 200, 0.2)',
                    showticklabels=False
                )
                
                fig1.update_xaxes(gridcolor='rgba(200, 200, 200, 0.2)')
                
                st.plotly_chart(fig1, use_container_width=True)
    
    # Seção de Avisos e Erros
    st.markdown("---")
    st.markdown("#### ⚠️ Avisos e Erros")
    st.info("⏳ **Esta seção está em construção.**")

def pagina_relatorios_conferencia_sinisa():
    """
    Relatório Conferência SINISA - Baseado no relatório Verificação
    Modificações:
    1. Remove filtro de Município
    2. Mantém filtro de Informação/Indicador
    3. Mantém subseção "📖 Detalhes da Informação"
    4. Mantém filtros e botão baixar em "📊 Composição da Informação"
       - Linhas: Municípios
       - Colunas: Componentes/Termos da informação selecionada
       - Valores: valor_acumulado do último mês com dados do ano
    5. Remove subseções "📈 Gráfico de Evolução" e "⚠️ Avisos e Erros"
    6. Adiciona aba com valores mensais de todos os municípios
    """
    st.markdown("""
    <div class="header-container">
        <h1 class="header-title">📚 SINISA</h1>
        <p class="header-subtitle">Sistema Nacional de Informações sobre Saneamento</p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("### 📋 Conferência SINISA")
    
    # Carregar dados - CORRIGIDO
    df_glossario = load_data()
    df_correspondencia = load_correspondencia_sigis_sinisa()
    df_dados_sigis = load_dados_sigis()
    df_sinisa = load_dados_sinisa()
    
    if df_glossario is None or df_correspondencia is None or df_dados_sigis is None or df_sinisa is None:
        st.error("❌ Não foi possível carregar os dados necessários")
        return
    
    # TRATAMENTO ROBUSTO DE DADOS
    df_correspondencia['codigo_sinisa'] = df_correspondencia['codigo_sinisa'].astype(str).str.strip()
    df_correspondencia['termos_sinisa'] = df_correspondencia['termos_sinisa'].astype(str).str.strip()
    
    def limpar_codigo_sigis(val):
        if pd.isna(val) or val == "" or str(val).strip() == "nan":
            return ""
        val_str = str(val).strip()
        if val_str.endswith('.0'):
            return val_str[:-2]
        return val_str
        
    df_correspondencia['termos_sigis'] = df_correspondencia['termos_sigis'].apply(limpar_codigo_sigis)
    df_dados_sigis['cod_sigis'] = df_dados_sigis['cod_sigis'].apply(limpar_codigo_sigis)
    
    # ========== FUNÇÕES AUXILIARES ==========
    
    @st.cache_data
    def get_informacoes_conferencia():
        """Obtém lista de informações únicas do glossário com código + descrição"""
        informacoes_dict = {}
        for _, row in df_glossario.iterrows():
            if pd.notna(row['Código da informação']):
                codigo = str(row['Código da informação']).strip()
                nome = str(row['Nome da informação']).strip()
                informacoes_dict[f"{codigo} - {nome}"] = codigo
        
        opcoes = ["Selecione uma informação"] + sorted(informacoes_dict.keys())
        return opcoes, informacoes_dict
    
    def get_descritivo_informacao(codigo_informacao):
        """Obtém descritivo completo da informação"""
        df_info = df_glossario[df_glossario['Código da informação'].astype(str).str.strip() == codigo_informacao]
        if len(df_info) == 0:
            return None
        return df_info.iloc[0]
    
    def get_componentes_informacao(codigo_selecionado):
        """Obtém os componentes (termos) de uma informação SINISA"""
        # 1. Tentar buscar como código principal (ex: GFI1003)
        df_comp_principal = df_correspondencia[df_correspondencia['codigo_sinisa'] == codigo_selecionado]
        
        if len(df_comp_principal) > 0:
            return df_comp_principal[['termos_sinisa', 'termos_sigis']].drop_duplicates().values.tolist()
            
        # 2. Se não encontrou, tentar buscar como termo (ex: GTA1209)
        df_comp_termo = df_correspondencia[df_correspondencia['termos_sinisa'] == codigo_selecionado]
        
        if len(df_comp_termo) > 0:
            return df_comp_termo[['termos_sinisa', 'termos_sigis']].drop_duplicates().values.tolist()
            
        return []
    
    def get_descricao_termo(termo_sinisa):
        """Obtém descrição de um termo SINISA"""
        df_termo = df_glossario[df_glossario['Código da informação'].astype(str).str.strip() == termo_sinisa]
        if len(df_termo) == 0:
            return ""
        return str(df_termo.iloc[0]['Nome da informação']).strip()
    
    # ========== SEÇÃO 1: FILTROS ==========
    st.markdown("#### 🔍 Filtros")
    
    # Obter opções de informações
    opcoes_info, info_dict = get_informacoes_conferencia()
    
    informacao_selecionada_display = st.selectbox(
        "Selecione a Informação/Indicador",
        opcoes_info,
        key="conferencia_sinisa_info",
        index=0
    )
    
    # Extrair código da seleção
    if informacao_selecionada_display != "Selecione uma informação":
        informacao_selecionada = info_dict[informacao_selecionada_display]
    else:
        informacao_selecionada = None
    
    if informacao_selecionada is None:
        st.warning("⚠️ Selecione uma informação para continuar.")
        return
    
    # Obter os componentes/termos da informação selecionada
    componentes = get_componentes_informacao(informacao_selecionada)
    
    if not componentes:
        st.warning(f"Nenhum componente encontrado para '{informacao_selecionada}'.")
        return
    
    # ========== SEÇÃO 2: DETALHES DA INFORMAÇÃO ==========
    st.markdown("#### 📖 Detalhes da Informação")
    
    row = get_descritivo_informacao(informacao_selecionada)
    
    if row is not None:
        exibir_detalhes_modal(row, "conferencia_0")
    
    st.markdown("---")
    
    # ========== SEÇÃO 3: COMPOSIÇÃO DA INFORMAÇÃO ==========
    st.markdown("#### 📊 Composição da Informação")
    
    # Filtro de Ano, Mensagem de Data e Botão Baixar - TUDO NA MESMA LINHA
    col_ano, col_data, col_botao = st.columns([0.8, 1.8, 0.8], vertical_alignment="center")
    
    with col_ano:
        # Filtro de Ano
        anos_disponiveis = sorted(df_dados_sigis['data'].dt.year.unique(), reverse=True)
        ano_selecionado = st.selectbox(
            "Ano",
            options=anos_disponiveis,
            key="conferencia_sinisa_ano",
            label_visibility="collapsed"
        )
    
    # Filtrar dados SIGIS pelo ano
    df_municipio = df_dados_sigis[
        (df_dados_sigis['data'].dt.year == ano_selecionado)
    ].copy()
    
    if len(df_municipio) == 0:
        st.warning(f"Nenhum dado disponível para {ano_selecionado}.")
        return
    
    # Converter data para datetime
    df_municipio['data'] = pd.to_datetime(df_municipio['data'])
    
    # Obter o último mês com dados do ano
    meses_disponiveis = sorted(df_municipio['data'].dt.month.unique())
    if not meses_disponiveis:
        st.warning(f"Nenhum dado disponível para {ano_selecionado}.")
        return
    
    ultimo_mes = meses_disponiveis[-1]
    
    # Filtrar para o último mês
    df_ultimo_mes = df_municipio[df_municipio['data'].dt.month == ultimo_mes].copy()
    
    # Criar tabela com municípios nas linhas e componentes nas colunas
    # Extrair apenas os códigos SIGIS dos componentes
    codigos_sigis = [termo_sigis for termo_sinisa, termo_sigis in componentes if termo_sigis]
    
    # Filtrar dados para os códigos SIGIS
    df_sigis_filtrado = df_ultimo_mes[df_ultimo_mes['cod_sigis'].isin(codigos_sigis)].copy()
    
    if len(df_sigis_filtrado) == 0:
        st.warning(f"Nenhum dado disponível para os componentes de '{informacao_selecionada}' em {ano_selecionado}.")
        return
    
    # Criar tabela pivot: municípios (linhas) x componentes (colunas)
    tabela_conferencia = df_sigis_filtrado.pivot_table(
        index='nome_localidade',
        columns='cod_sigis',
        values='valor_acumulado',
        aggfunc='first'
    )
    
    # Reordenar colunas conforme a ordem dos componentes
    colunas_ordenadas = [col for col in codigos_sigis if col in tabela_conferencia.columns]
    tabela_conferencia = tabela_conferencia[colunas_ordenadas]
    
    # Renomear colunas com descrição - MANTÉM CÓDIGO + DESCRIÇÃO
    colunas_renomeadas = {}
    for cod_sigis in colunas_ordenadas:
        df_desc = df_sigis_filtrado[df_sigis_filtrado['cod_sigis'] == cod_sigis]
        desc = df_desc['desc_sigis'].values[0] if len(df_desc) > 0 else cod_sigis
        
        # Manter código + descrição (ex: "1 - VOLUME PRODUZIDO TOTAL")
        colunas_renomeadas[cod_sigis] = desc
    
    tabela_conferencia = tabela_conferencia.rename(columns=colunas_renomeadas)
    
    # Renomear índice de nome_localidade para Município
    tabela_conferencia.index.name = 'Município'
    
    # Formatar valores
    tabela_conferencia_formatada = tabela_conferencia.copy()
    for col in tabela_conferencia_formatada.columns:
        tabela_conferencia_formatada[col] = tabela_conferencia_formatada[col].apply(
            lambda x: f"{x:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if pd.notna(x) else '-'
        )
    
    # Informação sobre o período - POSICIONADA ANTES DO BOTÃO
    with col_data:
        st.markdown(f"""
        <div style="background-color: #EBF2FB; padding: 8px 12px; border-radius: 4px; border-left: 4px solid #0B3040; font-size: 13px; color: #333;">
            📅 Dados referentes ao mês {ultimo_mes:02d}/{ano_selecionado}
        </div>
        """, unsafe_allow_html=True)
    
    # Botão Baixar - POSICIONADO À DIREITA
    with col_botao:
        # Preparar dados para download - COM MÚLTIPLAS ABAS
        excel_buffer = BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # ========== ABA 1: CONFERÊNCIA (Acumulado) ==========
            tabela_conferencia.to_excel(
                writer,
                sheet_name='Conferência',
                startrow=0
            )
            
            # ========== ABA 2: VALORES MENSAIS ==========
            # Criar tabela com todos os meses do ano
            df_todos_meses = df_municipio[df_municipio['cod_sigis'].isin(codigos_sigis)].copy()
            
            # Criar estrutura: Município | Código | Descrição | Mês 01 | Mês 02 | ... | Mês 12 | Anual
            dados_mensais = []
            municipios_unicos = sorted(df_todos_meses['nome_localidade'].unique())
            
            for municipio in municipios_unicos:
                df_mun = df_todos_meses[df_todos_meses['nome_localidade'] == municipio]
                
                # Para cada código SIGIS, criar uma linha
                for cod_sigis in codigos_sigis:
                    df_sigis_mun = df_mun[df_mun['cod_sigis'] == cod_sigis]
                    
                    if len(df_sigis_mun) > 0:
                        # Obter descrição do SIGIS
                        desc_sigis = df_sigis_mun['desc_sigis'].values[0]
                        
                        # Separar código e descrição
                        if '-' in desc_sigis:
                            partes = desc_sigis.split('-', 1)
                            codigo = partes[0].strip()
                            descricao = partes[1].strip()
                        else:
                            codigo = desc_sigis
                            descricao = ""
                        
                        linha = {
                            'Município': municipio,
                            'Código': codigo,
                            'Descrição': descricao
                        }
                        
                        # Adicionar valores de cada mês
                        for mes in range(1, 13):
                            df_mes = df_sigis_mun[df_sigis_mun['data'].dt.month == mes]
                            valor = df_mes['valor'].values[0] if len(df_mes) > 0 else 0
                            linha[f'{mes:02d}/{ano_selecionado}'] = valor
                        
                        # Adicionar valor anual (acumulado de dezembro)
                        df_dezembro = df_sigis_mun[df_sigis_mun['data'].dt.month == 12]
                        valor_anual = df_dezembro['valor_acumulado'].values[0] if len(df_dezembro) > 0 else 0
                        linha['Anual'] = valor_anual
                        
                        dados_mensais.append(linha)
            
            df_mensais = pd.DataFrame(dados_mensais)
            df_mensais.to_excel(
                writer,
                sheet_name='Valores Mensais',
                startrow=0,
                index=False
            )
            
            # Aplicar formatação em ambas as abas
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            workbook = writer.book
            
            # Formatação ABA 1: Conferência
            worksheet1 = writer.sheets['Conferência']
            header_fill = PatternFill(start_color="0B3040", end_color="0B3040", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin', color='d0d8e0'),
                right=Side(style='thin', color='d0d8e0'),
                top=Side(style='thin', color='d0d8e0'),
                bottom=Side(style='thin', color='d0d8e0')
            )
            
            # Aplicar estilo ao cabeçalho ABA 1
            for cell in worksheet1[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Aplicar formatação aos dados ABA 1
            for row in worksheet1.iter_rows(min_row=2, max_row=worksheet1.max_row, min_col=2, max_col=worksheet1.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0.00'
            
            # Ajustar largura das colunas ABA 1
            
             # Alinhar coluna Município para esquerda
            for row in worksheet1.iter_rows(min_row=1, max_row=worksheet1.max_row, min_col=1, max_col=1):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            for column in worksheet1.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet1.column_dimensions[column_letter].width = adjusted_width
            
            # Congelar primeira linha e primeira coluna ABA 1
            worksheet1.freeze_panes = 'B2'
            
            # Formatação ABA 2: Valores Mensais
            worksheet2 = writer.sheets['Valores Mensais']
            
            # Aplicar estilo ao cabeçalho ABA 2
            for cell in worksheet2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            
            # Aplicar formatação aos dados ABA 2
            # Colunas de valores (a partir da coluna 4 - Meses)
            for row in worksheet2.iter_rows(min_row=2, max_row=worksheet2.max_row, min_col=4, max_col=worksheet2.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0.00'
            
            # Aplicar formatação às colunas de Município, Código e Descrição (sem número)
            for row in worksheet2.iter_rows(min_row=2, max_row=worksheet2.max_row, min_col=1, max_col=3):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            
            # Ajustar largura das colunas ABA 2
            worksheet2.column_dimensions['A'].width = 25  # Município
            worksheet2.column_dimensions['B'].width = 10  # Código
            worksheet2.column_dimensions['C'].width = 40  # Descrição
            
            for col_idx in range(4, worksheet2.max_column + 1):
                column_letter = chr(64 + col_idx) if col_idx < 27 else chr(64 + col_idx // 26) + chr(64 + col_idx % 26)
                worksheet2.column_dimensions[column_letter].width = 15
            
            # Congelar primeira linha e primeiras três colunas ABA 2
            worksheet2.freeze_panes = 'D2'
        
        excel_buffer.seek(0)
        
        st.download_button(
            label="📥 Baixar Relatório",
            data=excel_buffer,
            file_name=f"Conferencia_SINISA_{informacao_selecionada}_{ano_selecionado}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="conferencia_sinisa_download",
            use_container_width=True
        )
    
    st.markdown("---")
    
    # Exibir tabela - CORRIGIDO: altura aumentada para mostrar todos os municípios
    num_municipios = len(tabela_conferencia_formatada)
    altura_tabela = max(400, num_municipios * 35 + 50)  # 35px por linha + 50px para cabeçalho
    
    st.dataframe(
        tabela_conferencia_formatada,
        use_container_width=True,
        height=altura_tabela
    )
def pagina_infraestrutura():
    """Página Infraestrutura"""
    st.markdown("""
    <div class="header-container">
        <h1 class="header-title">📚 SINISA</h1>
        <p class="header-subtitle">Sistema Nacional de Informações sobre Saneamento</p>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("### 🏗️ Infraestrutura")
    
    st.info("""
    ⏳ **Esta seção está em construção.**
    
    Em breve, você poderá acessar informações sobre a infraestrutura e configurações do sistema SINISA.
    """)

def pagina_relatorios_indicadores_ana():
    """Página Relatórios - Indicadores ANA"""
    st.markdown("""
    <div class="header-container">
        <h1 class="header-title">📚 SINISA</h1>
        <p class="header-subtitle">Sistema Nacional de Informações sobre Saneamento</p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("### 💧 Indicadores ANA")
    st.info("""
    ⏳ **Esta seção está em fase de construção.**
    Em breve, você poderá acessar os relatórios de Indicadores da ANA.
    """)

def pagina_relatorios_rad_agems():
    """Página Relatórios - RAD (AGEMS)"""
    st.markdown("""
    <div class="header-container">
        <h1 class="header-title">📚 SINISA</h1>
        <p class="header-subtitle">Sistema Nacional de Informações sobre Saneamento</p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("### 📈 RAD (AGEMS)")
    st.info("""
    ⏳ **Esta seção está em fase de construção.**
    Em breve, você poderá acessar os relatórios RAD da AGEMS.
    """)

def pagina_contrato_programa():
    """Página Contrato Programa"""
    st.markdown("""
    <div class="header-container">
        <h1 class="header-title">📚 SINISA</h1>
        <p class="header-subtitle">Sistema Nacional de Informações sobre Saneamento</p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("### 📄 Contrato Programa")
    st.info("""
    ⏳ **Esta seção está em fase de construção.**
    Em breve, você poderá acessar as informações sobre Contratos de Programa.
    """)

def pagina_investimento():
    """Página Investimento"""
    st.markdown("""
    <div class="header-container">
        <h1 class="header-title">📚 SINISA</h1>
        <p class="header-subtitle">Sistema Nacional de Informações sobre Saneamento</p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("### 💰 Investimento")
    st.info("""
    ⏳ **Esta seção está em fase de construção.**
    Em breve, você poderá acessar os dados e relatórios de Investimentos.
    """)    

# ============================================================================
# ESTILOS GLOBAIS PARA SIDEBAR - VERSÃO FINAL OTIMIZADA
# ============================================================================

SIDEBAR_STYLES = """
<style>
    /* ========== SIDEBAR CONTAINER ========== */
    [data-testid="stSidebar"] {
        background-color: #f5f7fa !important;
        padding: 0 !important;
    }
    
    /* ========== SIDEBAR CONTENT ========== */
    [data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
        padding: 1rem !important;
        gap: 0.5rem !important;
    }
    
    /* ========== TÍTULOS (H3) ========== */
    [data-testid="stSidebar"] h3 {
        color: #1f4788 !important;
        font-size: 0.9rem !important;
        font-weight: 600 !important;
        margin: 0.75rem 0 0.5rem 0 !important;
        padding: 0 !important;
        text-align: left !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
    }
    
    /* ========== EXPANDERS (TÓPICOS) ========== */
    [data-testid="stSidebar"] .streamlit-expanderHeader {
        background-color: #e8eef5 !important;
        border: 1px solid #d0d8e0 !important;
        border-radius: 6px !important;
        padding: 0.5rem 0.8rem !important;
        margin-bottom: 0.5rem !important;
        font-weight: 600 !important;
        font-size: 0.85rem !important;
        color: #1f4788 !important;
        text-align: left !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
    }
    
    [data-testid="stSidebar"] .streamlit-expanderHeader:hover {
        background-color: #d4e4f7 !important;
        border-color: #2d5aa8 !important;
    }
    
    [data-testid="stSidebar"] .streamlit-expanderHeader p {
        margin: 0 !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        font-size: 0.85rem !important;
        text-align: left !important;
    }
    
    /* ========== CONTEÚDO DENTRO DO EXPANDER ========== */
    [data-testid="stSidebar"] .streamlit-expanderContent {
        padding: 0.5rem 0 !important;
        border-left: 3px solid #2d5aa8 !important;
        margin-left: 0.5rem !important;
    }
    
    /* ========== BOTÕES NO SIDEBAR ========== */
    [data-testid="stSidebar"] button {
        width: 100% !important;
        text-align: left !important;
        padding: 0.5rem 0.8rem !important;
        margin: 0.25rem 0 !important;
        border: 1px solid #d0d8e0 !important;
        border-radius: 4px !important;
        background-color: #ffffff !important;
        color: #333 !important;
        font-size: 0.85rem !important;
        font-weight: 500 !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        transition: all 0.2s ease !important;
        display: flex !important;
        align-items: center !important;
        justify-content: flex-start !important;
    }
    
    [data-testid="stSidebar"] button:hover {
        background-color: #e8eef5 !important;
        border-color: #2d5aa8 !important;
        color: #1f4788 !important;
    }
    
    [data-testid="stSidebar"] button:active {
        background-color: #d4e4f7 !important;
        border-color: #1f4788 !important;
    }
    
    /* ========== TEXTO DENTRO DOS BOTÕES ========== */
    [data-testid="stSidebar"] button p {
        margin: 0 !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        font-size: 0.85rem !important;
        text-align: left !important;
    }
    
    /* ========== ÍCONES DENTRO DOS BOTÕES ========== */
    [data-testid="stSidebar"] button svg {
        margin-right: 0.5rem !important;
        flex-shrink: 0 !important;
    }
    
    /* ========== MARKDOWN NO SIDEBAR ========== */
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] {
        text-align: left !important;
    }
    
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p {
        margin: 0 !important;
        font-size: 0.85rem !important;
        text-align: left !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
    }
    
    /* ========== ESPAÇAMENTO GERAL ========== */
    [data-testid="stSidebar"] [data-testid="stVerticalBlock"] > div {
        margin: 0 !important;
    }
    
    /* ========== REMOVER ESPAÇOS EXTRAS ========== */
    [data-testid="stSidebar"] [data-testid="stVerticalBlock"] > div > div {
        margin: 0 !important;
        padding: 0 !important;
    }
</style>
"""

st.markdown(SIDEBAR_STYLES, unsafe_allow_html=True)

# ============================================================================
# DICIONÁRIO DE PÁGINAS
# ============================================================================
PAGES = {
    "Sobre": pagina_sobre,
    "Glossário": {
        "Informações": pagina_glossario_informacoes_com_stats,
        "Indicadores": pagina_glossario_indicadores,
        "Avisos e Erros": pagina_glossario_avisos,

    },
    "Relatórios": {
        "Gerencial por Município": pagina_relatorios_municipio,
        "Consolidado": pagina_relatorios_consolidado,
        "Verificação": pagina_relatorios_verificacao,
        "Conferência SINISA": pagina_relatorios_conferencia_sinisa,
        "Indicadores ANA": pagina_relatorios_indicadores_ana,
        "RAD (AGEMS)": pagina_relatorios_rad_agems,

    },
    "Contrato Programa": pagina_contrato_programa,
    "Investimento": pagina_investimento,
    "Infraestrutura": pagina_infraestrutura,
}

# ============================================================================
# NAVEGAÇÃO NO SIDEBAR
# ============================================================================

def render_sidebar_navigation():
    """Renderiza navegação no sidebar com melhor organização"""
    with st.sidebar:
        st.markdown("### 📌 Navegação")
        
        # INICIALIZAÇÃO SEGURA DO SESSION STATE (Evita o KeyError)
        if "current_page" not in st.session_state:
            st.session_state["current_page"] = "Sobre"
            
        # Página Sobre
        if st.button("📖 Sobre", use_container_width=True, key="btn_sobre"):
            st.session_state["current_page"] = "Sobre"
            st.rerun()
            
        # Glossário (com subtópicos)
        with st.expander("📚 Glossário", expanded=False):
            if st.button("📋 Informações", use_container_width=True, key="btn_info"):
                st.session_state["current_page"] = "Informações"
                st.rerun()
            if st.button("📈 Indicadores", use_container_width=True, key="btn_ind"):
                st.session_state["current_page"] = "Indicadores"
                st.rerun()
            if st.button("⚠️ Avisos e Erros", use_container_width=True, key="btn_avisos"):
                st.session_state["current_page"] = "Avisos e Erros"
                st.rerun()
                
        # Relatórios (com subtópicos)
        with st.expander("📊 Relatórios", expanded=False):
            if st.button("🏘️ Gerencial por Município", use_container_width=True, key="btn_mun"):
                st.session_state["current_page"] = "Gerencial por Município"
                st.rerun()
            if st.button("📊 Consolidado", use_container_width=True, key="btn_cons"):
                st.session_state["current_page"] = "Consolidado"
                st.rerun()

            if st.button("✅ Verificação", use_container_width=True, key="btn_verif"):
                st.session_state["current_page"] = "Verificação"
                st.rerun()

            if st.button("☑️ Conferência SINISA", use_container_width=True, key="btn_confSinisa"):
                st.session_state["current_page"] = "Conferência SINISA"
                st.rerun()

            if st.button("💧 Indicadores ANA", use_container_width=True, key="btn_ana"):
                st.session_state["current_page"] = "Indicadores ANA"
                st.rerun()

            if st.button("📈 RAD (AGEMS)", use_container_width=True, key="btn_rad"):
                st.session_state["current_page"] = "RAD (AGEMS)"
                st.rerun()
                
        # Infraestrutura
        if st.button("🏗️ Infraestrutura", use_container_width=True, key="btn_infra"):
            st.session_state["current_page"] = "Infraestrutura"
            st.rerun()

        # Contrato Programa
        if st.button("📄 Contrato Programa", use_container_width=True, key="btn_contrato"):
            st.session_state["current_page"] = "Contrato Programa"
            st.rerun()

        # Investimento
        if st.button("💰 Investimento", use_container_width=True, key="btn_investimento"):
            st.session_state["current_page"] = "Investimento"
            st.rerun()
                


def get_current_page_function():
    """Obtém a função da página atual"""
    current = st.session_state.get("current_page", "Sobre")
    
    # Procurar em páginas únicas
    if current in PAGES and not isinstance(PAGES[current], dict):
        return PAGES[current]
    
    # Procurar em subtópicos
    for section_name, section_content in PAGES.items():
        if isinstance(section_content, dict):
            if current in section_content:
                return section_content[current]
    
    # Padrão: retornar página "Sobre"
    return PAGES["Sobre"]

# ============================================================================
# EXECUÇÃO PRINCIPAL
# ============================================================================

def main():
    """Função principal"""
    # Renderizar sidebar
    render_sidebar_navigation()
    
    # Executar página atual
    current_func = get_current_page_function()
    current_func()

if __name__ == "__main__":
    main()
